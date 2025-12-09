import io
import csv
import os
import json
from pathlib import Path
from functools import wraps
from flask import Flask, render_template, request, jsonify, send_file, session, redirect, url_for, flash
from flask_sqlalchemy import SQLAlchemy
from datetime import datetime, timedelta
import random
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from werkzeug.security import generate_password_hash, check_password_hash
import ssl
import tempfile
import os
import stat
import psycopg2
import re
import os
import stat
import logging

def verificar_certificados():
    """Verifica se os certificados existem e estão corretos"""
    certificados = ['ca-certificate.crt', 'certificate.pem', 'private-key.key']
    
    for cert in certificados:
        if not os.path.exists(cert):
            print(f"❌ Certificado não encontrado: {cert}")
            return False
        else:
            print(f"✅ Certificado encontrado: {cert}")
    
    return True

def corrigir_permissoes_certificados():
    """Corrige permissões dos certificados PostgreSQL"""
    certificados = {
        'private-key.key': stat.S_IRUSR | stat.S_IWUSR,  # 600
        'certificate.pem': stat.S_IRUSR | stat.S_IWUSR,  # 600
        'ca-certificate.crt': stat.S_IRUSR | stat.S_IWUSR | stat.S_IRGRP | stat.S_IROTH  # 644
    }
    
    for cert_file, perms in certificados.items():
        if os.path.exists(cert_file):
            try:
                os.chmod(cert_file, perms)
                print(f"✅ Permissões ajustadas para: {cert_file}")
            except Exception as e:
                print(f"⚠️ Não foi possível ajustar {cert_file}: {e}")

# Chame esta função ANTES de db.create_all()
corrigir_permissoes_certificados()


# CORREÇÃO DO FUSO HORÁRIO
def agora():
    """Retorna o horário atual de Brasília (UTC-3)"""
    return datetime.utcnow() - timedelta(hours=3)

# Configurações PostgreSQL COM CERTIFICADOS
# app.py - na seção da classe Config

class Config:
    SECRET_KEY = os.environ.get('SECRET_KEY') or 'sua-chave-secreta-super-segura-aqui-ro-experience-2025'
    
    # 1. URI DE CONEXÃO CORRIGIDA E ATUALIZADA:
    # - Usa o driver +psycopg2
    # - Usa a senha nova: RGil4Y8VE9M5znboTfwHInhm
    # - Aponta para o DB final (dbexperience)
    # - Inclui TODOS os parâmetros SSL na query string (obrigatórios pelo servidor)
    SQLALCHEMY_DATABASE_URI = os.environ.get('DATABASE_URL') or \
        'postgresql+psycopg2://squarecloud:RGil4Y8VE9M5znboTfwHInhm@square-cloud-db-4d0ca60ac1a54ad48adf5608996c6a48.squareweb.app:7091/dbexperience?sslmode=require&sslrootcert=ca-certificate.crt&sslcert=certificate.pem&sslkey=private-key.key'
    
    # 2. REMOVA SQLALCHEMY_ENGINE_OPTIONS:
    # Este bloco é desnecessário e causa conflitos quando o SSL está na URI.
    # Certifique-se de que ele não existe mais.
    # SQLALCHEMY_ENGINE_OPTIONS = { ... } <-- REMOVA!
    
    SQLALCHEMY_TRACK_MODIFICATIONS = False

app = Flask(__name__)
app.config.from_object(Config)
db = SQLAlchemy(app)

# CONFIGURAÇÃO DE LOGGING
logging.basicConfig(level=logging.DEBUG)
logger = logging.getLogger(__name__)

# Modelos

# NOVO MODELO PARA VENDAS DO EVENTO (DADOS HISTÓRICOS)
# NOVO MODELO COMPLETO - ADICIONE O CAMPO numero_nf
class VendaEvento(db.Model):
    __tablename__ = 'venda_evento'
    __table_args__ = {'extend_existing': True}  # ← ADICIONE ESTA LINHA
    
    id = db.Column(db.Integer, primary_key=True)
    numero_nf = db.Column(db.String(50), nullable=True)  # ← OBRIGATÓRIO!
    data_emissao = db.Column(db.Date, nullable=False)
    cliente_nome = db.Column(db.String(200), nullable=False)
    vendedor = db.Column(db.String(100), nullable=False)
    equipe = db.Column(db.String(100), nullable=False)
    descricao_produto = db.Column(db.String(300), nullable=False)
    marca = db.Column(db.String(100), nullable=False)
    valor_produtos = db.Column(db.Float, nullable=False)
    quantidade = db.Column(db.Integer, nullable=False, default=1)
    familia = db.Column(db.String(100))
    valor_total = db.Column(db.Float, nullable=False)
    data_importacao = db.Column(db.DateTime, default=agora)
    importado_por = db.Column(db.String(100))
    
    # Índices
    __table_args__ = (
        db.Index('idx_venda_evento_cliente', 'cliente_nome'),
        db.Index('idx_venda_evento_vendedor', 'vendedor'),
        db.Index('idx_venda_evento_data', 'data_emissao'),
        db.Index('idx_venda_evento_marca', 'marca'),
        db.Index('idx_venda_evento_familia', 'familia'),
        db.Index('idx_venda_evento_nf', 'numero_nf'),
    )
    
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        # Garante que valor_total seja igual a valor_produtos
        if self.valor_total is None:
            self.valor_total = self.valor_produtos

class Usuario(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False)
    password_hash = db.Column(db.String(200), nullable=False)
    nome = db.Column(db.String(100), nullable=False)
    nivel_acesso = db.Column(db.String(20), default='operador')
    ativo = db.Column(db.Boolean, default=True)
    data_criacao = db.Column(db.DateTime, default=agora)
    permissoes = db.Column(db.Text, default='{}')

class LogAuditoria(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    usuario_id = db.Column(db.Integer, db.ForeignKey('usuario.id'))
    acao = db.Column(db.String(100), nullable=False)
    modulo = db.Column(db.String(50), nullable=False)
    dados = db.Column(db.Text)
    ip = db.Column(db.String(45))
    data_hora = db.Column(db.DateTime, default=agora)
    
    usuario = db.relationship('Usuario', backref='logs')

class Cliente(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    cnpj = db.Column(db.String(18), nullable=False)
    razao_social = db.Column(db.String(200), nullable=False)
    responsavel = db.Column(db.String(200), nullable=False, default='Não Informado')
    consultor = db.Column(db.String(200), nullable=False, default='Não Informado')
    checkin_realizado = db.Column(db.Boolean, default=False)
    horario_checkin = db.Column(db.DateTime)
    responsavel_checkin = db.Column(db.String(200))
    direito_imagem = db.Column(db.Boolean, default=False)
    veio_carro = db.Column(db.Boolean, default=False)
    placa_veiculo = db.Column(db.String(10))
    
    vendas = db.relationship('Venda', backref='cliente', lazy=True)

    __table_args__ = (db.UniqueConstraint('cnpj', 'responsavel', name='uq_cnpj_responsavel'),)

class Venda(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    cnpj_compra = db.Column(db.String(18), nullable=False)
    numero_pedido = db.Column(db.String(50), unique=True, nullable=False)
    valor_pedido = db.Column(db.Float, nullable=False)
    data_hora_venda = db.Column(db.DateTime, default=agora)
    cliente_id = db.Column(db.Integer, db.ForeignKey('cliente.id'))
    cnpj_checkin_vinculado = db.Column(db.String(18))
    
    equipamentos = db.relationship('VendaEquipamento', backref='venda', lazy=True)

class Brinde(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    tipo_sorteio = db.Column(db.String(10), nullable=False)
    nome = db.Column(db.String(200), nullable=False)
    descricao = db.Column(db.Text)
    valor_aproximado = db.Column(db.Float)
    quantidade_total = db.Column(db.Integer, nullable=False, default=1)
    quantidade_disponivel = db.Column(db.Integer, nullable=False, default=1)
    ativo = db.Column(db.Boolean, default=True)

class Sorteio(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    tipo_brinde = db.Column(db.String(50), nullable=False)
    cnpj_vencedor = db.Column(db.String(18), nullable=False)
    razao_social_vencedor = db.Column(db.String(200), nullable=False)
    responsavel_recebimento = db.Column(db.String(200), nullable=False)
    data_sorteio = db.Column(db.DateTime, default=agora)
    valor_acumulado_revenda = db.Column(db.Float, nullable=False)
    brinde_id = db.Column(db.Integer, db.ForeignKey('brinde.id'))
    brinde = db.relationship('Brinde', backref='sorteios')
    entregue = db.Column(db.Boolean, default=False)
    data_entrega = db.Column(db.DateTime)
    responsavel_entrega = db.Column(db.String(200))
    observacao_entrega = db.Column(db.Text)

class FaturamentoSorteio(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    cnpj = db.Column(db.String(18), nullable=False)
    faturamento_acumulado = db.Column(db.Float, nullable=False, default=0.0)
    ultima_atualizacao = db.Column(db.DateTime, default=agora)
    participacoes_utilizadas = db.Column(db.Integer, default=0)

class Estoque(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    fabricante = db.Column(db.String(100), nullable=False)
    modelo = db.Column(db.String(100), nullable=False)
    quantidade_total = db.Column(db.Integer, nullable=False, default=0)
    quantidade_disponivel = db.Column(db.Integer, nullable=False, default=0)
    data_cadastro = db.Column(db.DateTime, default=agora)
    ativo = db.Column(db.Boolean, default=True)
    
    vendas = db.relationship('VendaEquipamento', backref='equipamento', lazy=True)

class VendaEquipamento(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    venda_id = db.Column(db.Integer, db.ForeignKey('venda.id'))
    equipamento_id = db.Column(db.Integer, db.ForeignKey('estoque.id'))
    quantidade = db.Column(db.Integer, nullable=False, default=1)

class PesquisaResposta(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    cnpj_identificado = db.Column(db.String(18))
    razao_social = db.Column(db.String(200))
    
    # NOVAS PERGUNTAS (1-5)
    comunicacao = db.Column(db.Integer, nullable=False)  # 1-5
    formato_evento = db.Column(db.Integer, nullable=False)  # 1-5
    alimentacao = db.Column(db.Integer, nullable=False)  # 1-5
    palestra_reforma = db.Column(db.Integer, nullable=False)  # 1-5
    palestra_estrategia = db.Column(db.Integer, nullable=False)  # 1-5
    
    # PERGUNTAS EXISTENTES ATUALIZADAS (6-15)
    organizacao = db.Column(db.Integer, nullable=False)  # 1-5
    interacao_brother = db.Column(db.Integer, nullable=False)  # 1-5
    interacao_canon = db.Column(db.Integer, nullable=False)  # 1-5
    interacao_epson = db.Column(db.Integer, nullable=False)  # 1-5
    interacao_hp = db.Column(db.Integer, nullable=False)  # 1-5
    interacao_konica = db.Column(db.Integer, nullable=False)  # 1-5
    interacao_kyocera = db.Column(db.Integer, nullable=False)  # 1-5
    prazo_entrega = db.Column(db.Integer, nullable=False)  # 1-5
    frete = db.Column(db.Integer, nullable=False)  # 1-5
    
    comentarios = db.Column(db.Text)
    data_resposta = db.Column(db.DateTime, default=agora)
    ip = db.Column(db.String(45))
    anonima = db.Column(db.Boolean, default=False)

class PesquisaMarketing(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    cnpj_identificado = db.Column(db.String(18))
    razao_social = db.Column(db.String(200))
    
    # Questão 1
    posicionamento = db.Column(db.String(20), nullable=False)
    
    # Questão 2 (múltipla escolha - até 2)
    beneficio_engajamento = db.Column(db.Text)  # JSON array
    
    # Questão 3
    beneficio_preferido = db.Column(db.String(100), nullable=False)
    
    # Questão 4 (avaliações 1-5) - TORNAR OPCIONAIS
    margem_lucro = db.Column(db.Integer, default=0)
    qualidade_produtos = db.Column(db.Integer, default=0)
    suporte_comercial = db.Column(db.Integer, default=0)
    condicoes_comerciais = db.Column(db.Integer, default=0)
    reconhecimento_marca = db.Column(db.Integer, default=0)
    velocidade_resposta = db.Column(db.Integer, default=0)
    facilidade_pedidos = db.Column(db.Integer, default=0)
    
    # Questões 5-10 - TORNAR OPCIONAIS
    dificuldade_participacao = db.Column(db.String(100))
    tipo_campanha_impacto = db.Column(db.String(100))
    beneficio_venda = db.Column(db.Text)
    aumento_volume = db.Column(db.String(100))
    competitividade = db.Column(db.String(20))
    valor_parceiro = db.Column(db.String(100))
    
    # Outros campos
    outro_questao2 = db.Column(db.String(200))
    outro_questao3 = db.Column(db.String(200))
    outro_questao5 = db.Column(db.String(200))
    outro_questao6 = db.Column(db.String(200))
    outro_questao8 = db.Column(db.String(200))
    outro_questao10 = db.Column(db.String(200))
    
    comentarios_gerais = db.Column(db.Text)
    data_resposta = db.Column(db.DateTime, default=agora)
    ip = db.Column(db.String(45))
    usuario_responsavel = db.Column(db.String(200))

MODULOS_SISTEMA = {
    'dashboard': {'nome': '📊 Dashboard', 'descricao': 'Página inicial do sistema'},
    'checkin': {'nome': '✅ Check-in', 'descricao': 'Check-in de participantes'},
    'vendas': {'nome': '💰 Vendas', 'descricao': 'Registro de vendas'},
    'estoque': {'nome': '📦 Estoque', 'descricao': 'Gestão de estoque'},
    'sorteio': {'nome': '🎁 Sorteio', 'descricao': 'Realizar sorteios de brindes'},
    'entrega_brindes': {'nome': '✅ Entrega Brindes', 'descricao': 'Confirmar entrega de brindes sorteados'},
    'relatorios': {'nome': '📈 Relatórios', 'descricao': 'Relatórios do sistema'},
    'importacao': {'nome': '📤 Importação', 'descricao': 'Importar dados'},
    'exportacao': {'nome': '📥 Exportação', 'descricao': 'Exportar dados'},
    'brindes': {'nome': '🎯 Brindes', 'descricao': 'Gestão de brindes'},
    'usuarios': {'nome': '👥 Usuários', 'descricao': 'Gestão de usuários'},
    'logs': {'nome': '📊 Logs', 'descricao': 'Logs de auditoria'},
    
    # MÓDULOS DE PESQUISA (COMPLETO)
    'pesquisa_publica': {'nome': '📝 Pesquisa Pública', 'descricao': 'Pesquisa de satisfação do evento'},
    'pesquisa_marketing': {'nome': '📈 Pesquisa Marketing', 'descricao': 'Pesquisa de estratégia comercial'},
    'relatorio_pesquisas': {'nome': '📊 Relatório Pesquisas', 'descricao': 'Relatórios das pesquisas de satisfação'},
    'relatorio_pesquisas_mkt': {'nome': '📈 Relatório Pesq. Marketing', 'descricao': 'Relatórios das pesquisas de marketing'},
    'analise_vendas': {'nome': '📊 Análise de Vendas', 'descricao': 'Análise detalhada das vendas do evento'},
    'importacao_vendas': {'nome': '📥 Importar Vendas', 'descricao': 'Importar dados históricos de vendas'},
   

}


# Funções de permissão
def permissao_required(modulo):
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if session.get('nivel_acesso') == 'admin':
                return f(*args, **kwargs)
            
            if not tem_permissao(modulo):
                flash(f'❌ Acesso negado. Permissão necessária para: {MODULOS_SISTEMA[modulo]["nome"]}', 'error')
                return redirect(url_for('index'))
            
            return f(*args, **kwargs)
        return decorated_function
    return decorator

def migrar_banco_dados():
    """Função de migração adaptada para PostgreSQL"""
    try:
        with db.engine.connect() as conn:
            result = conn.execute(db.text("""
                SELECT column_name 
                FROM information_schema.columns 
                WHERE table_name='faturamento_sorteio' AND column_name='participacoes_utilizadas'
            """))
            coluna_existe = result.fetchone() is not None
            
            if not coluna_existe:
                print("🔄 Adicionando campo 'participacoes_utilizadas' à tabela faturamento_sorteio...")
                conn.execute(db.text("ALTER TABLE faturamento_sorteio ADD COLUMN participacoes_utilizadas INTEGER DEFAULT 0"))
                conn.commit()
                print("✅ Campo 'participacoes_utilizadas' adicionado com sucesso!")
            else:
                print("✅ Campo 'participacoes_utilizadas' já existe na tabela")
    except Exception as e:
        print(f"❌ Erro na migração: {e}")

def tem_permissao(modulo):
    """Verifica se usuário tem permissão para o módulo"""
    if session.get('nivel_acesso') == 'admin':
        return True
    
    usuario = Usuario.query.get(session['usuario_id'])
    if usuario and usuario.permissoes:
        try:
            permissoes = json.loads(usuario.permissoes)
            return permissoes.get(modulo, False)
        except json.JSONDecodeError:
            return False
    
    return False

def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'usuario_id' not in session:
            flash('🔐 Por favor, faça login para acessar esta página', 'info')
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'usuario_id' not in session or session.get('nivel_acesso') != 'admin':
            flash('Acesso negado. Permissão de administrador necessária.', 'error')
            return redirect(url_for('dashboard'))
        return f(*args, **kwargs)
    return decorated_function

def criar_usuario_admin():
    """Cria usuário admin padrão se não existir"""
    admin = Usuario.query.filter_by(username='admin').first()
    if not admin:
        admin = Usuario(
            username='admin',
            password_hash=generate_password_hash('admin123'),
            nome='Administrador',
            nivel_acesso='admin'
        )
        db.session.add(admin)
        db.session.commit()
        print("✅ Usuário admin criado: admin / admin123")

def registrar_log(acao, modulo, dados=None):
    """Registra uma ação no log de auditoria"""
    log = LogAuditoria(
        usuario_id=session.get('usuario_id'),
        acao=acao,
        modulo=modulo,
        dados=json.dumps(dados, ensure_ascii=False) if dados else None,
        ip=request.remote_addr
    )
    db.session.add(log)
    db.session.commit()

# Funções auxiliares
def normalizar_cnpj(cnpj):
    if not cnpj:
        return ""
    cnpj_limpo = ''.join(filter(str.isdigit, str(cnpj)))
    if len(cnpj_limpo) == 14:
        return f"{cnpj_limpo[:2]}.{cnpj_limpo[2:5]}.{cnpj_limpo[5:8]}/{cnpj_limpo[8:12]}-{cnpj_limpo[12:]}"
    return cnpj_limpo

def normalizar_cnpj_pesquisa(cnpj):
    if not cnpj:
        return ""
    
    cnpj_limpo = ''.join(filter(str.isdigit, str(cnpj)))
    
    if len(cnpj_limpo) == 14:
        return f"{cnpj_limpo[:2]}.{cnpj_limpo[2:5]}.{cnpj_limpo[5:8]}/{cnpj_limpo[8:12]}-{cnpj_limpo[12:]}"
    else:
        return cnpj_limpo

def format_currency(value):
    if value is None or value == 0:
        return "0,00"
    try:
        value_float = float(value)
        return "{:,.2f}".format(value_float).replace(",", "X").replace(".", ",").replace("X", ".")
    except (ValueError, TypeError):
        return "0,00"

app.jinja_env.filters['currency'] = format_currency
app.jinja_env.filters['cnpj'] = normalizar_cnpj

def get_faturamento_para_sorteio(cnpj):
    faturamento = FaturamentoSorteio.query.filter_by(cnpj=cnpj).first()
    if faturamento:
        return faturamento.faturamento_acumulado
    return 0.0

def get_revendas_para_sorteio(tipo_brinde):
    revendas_unicas = db.session.query(
        Cliente.cnpj,
        Cliente.razao_social
    ).distinct().all()
    
    revendas_qualificadas = []
    
    for cnpj, razao_social in revendas_unicas:
        faturamento = get_faturamento_para_sorteio(cnpj)
        
        if tipo_brinde == '50k':
            participacoes_totais = int(faturamento // 50000) if faturamento >= 50000 else 0
            
            if participacoes_totais > 0:
                sorteios_realizados = Sorteio.query.filter_by(
                    cnpj_vencedor=cnpj, 
                    tipo_brinde='50k'
                ).count()
                
                participacoes_restantes = participacoes_totais - sorteios_realizados
                
                if participacoes_restantes > 0:
                    revendas_qualificadas.append({
                        'cnpj': cnpj,
                        'razao_social': razao_social,
                        'faturamento_total': faturamento,
                        'participacoes': participacoes_restantes,
                        'sorteios_realizados': sorteios_realizados
                    })
                    
        elif tipo_brinde == '20k' and faturamento >= 20000:
            sorteios_realizados = Sorteio.query.filter_by(
                cnpj_vencedor=cnpj, 
                tipo_brinde='20k'
            ).count()
            
            if sorteios_realizados == 0 and faturamento < 50000:
                revendas_qualificadas.append({
                    'cnpj': cnpj,
                    'razao_social': razao_social,
                    'faturamento_total': faturamento,
                    'participacoes': 1,
                    'sorteios_realizados': 0
                })
    
    return revendas_qualificadas

def atualizar_faturamento_sorteio():
    FaturamentoSorteio.query.delete()
    
    faturamento_por_cnpj = db.session.query(
        Cliente.cnpj,
        db.func.sum(Venda.valor_pedido).label('faturamento_total')
    ).join(Venda).group_by(Cliente.cnpj).all()
    
    for cnpj, faturamento_total in faturamento_por_cnpj:
        sorteios_realizados = Sorteio.query.filter_by(cnpj_vencedor=cnpj, tipo_brinde='50k').count()
        
        faturamento_entry = FaturamentoSorteio(
            cnpj=cnpj,
            faturamento_acumulado=float(faturamento_total) if faturamento_total else 0.0,
            participacoes_utilizadas=sorteios_realizados
        )
        db.session.add(faturamento_entry)
    
    db.session.commit()

def get_estatisticas_avancadas():
    total_clientes = Cliente.query.count()
    total_checkins = Cliente.query.filter_by(checkin_realizado=True).count()
    revendas_unicas = db.session.query(Cliente.cnpj).distinct().count()
    revendas_presentes = db.session.query(Cliente.cnpj)\
        .filter(Cliente.checkin_realizado == True)\
        .distinct()\
        .count()
    
    return {
        'total_clientes': total_clientes,
        'total_checkins': total_checkins,
        'revendas_unicas': revendas_unicas,
        'revendas_presentes': revendas_presentes
    }

def export_to_excel(data, filename, sheet_name="Dados"):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    
    if data:
        headers = list(data[0].keys())
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        
        for row_num, row_data in enumerate(data, 2):
            for col_num, key in enumerate(headers, 1):
                ws.cell(row=row_num, column=col_num, value=row_data.get(key, ''))
    
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output

def read_excel_file(file_stream):
    wb = load_workbook(filename=file_stream)
    ws = wb.active
    
    data = []
    headers = []
    
    for row_num, row in enumerate(ws.iter_rows(values_only=True), 1):
        if row_num == 1:
            headers = [str(header).strip().upper() if header else '' for header in row]
        else:
            row_data = {}
            for col_num, value in enumerate(row):
                if col_num < len(headers):
                    row_data[headers[col_num]] = value
            if any(row_data.values()):
                data.append(row_data)
    
    return data

def get_participacoes_50k(cnpj):
    faturamento = get_faturamento_para_sorteio(cnpj)
    if faturamento < 50000:
        return 0
    
    participacoes_totais = int(faturamento // 50000)
    sorteios_realizados = Sorteio.query.filter_by(cnpj_vencedor=cnpj, tipo_brinde='50k').count()
    
    return max(0, participacoes_totais - sorteios_realizados)

# Rotas principais
@app.route('/')
def index():
    if 'usuario_id' not in session:
        return redirect(url_for('login'))
    
    if tem_permissao('dashboard'):
        total_vendas = db.session.query(db.func.sum(Venda.valor_pedido)).scalar() or 0
        estatisticas = get_estatisticas_avancadas()
        consultores_count = db.session.query(Cliente.consultor).distinct().count()
        
        return render_template('index.html',
                            total_vendas=total_vendas,
                            total_checkins=estatisticas['total_checkins'],
                            total_clientes=estatisticas['total_clientes'],
                            revendas_unicas=estatisticas['revendas_unicas'],
                            revendas_presentes=estatisticas['revendas_presentes'],
                            consultores_count=consultores_count,
                            now=agora())
    else:
        return render_template('index.html', now=agora())

@app.route('/dashboard')
@login_required
@permissao_required('dashboard')
def dashboard():
    total_vendas = db.session.query(db.func.sum(Venda.valor_pedido)).scalar() or 0
    estatisticas = get_estatisticas_avancadas()
    consultores_count = db.session.query(Cliente.consultor).distinct().count()
    
    vendas_por_revenda = db.session.query(
        Cliente.razao_social,
        db.func.sum(Venda.valor_pedido).label('total_vendas')
    ).join(Venda).group_by(Cliente.razao_social).order_by(db.desc('total_vendas')).limit(10).all()
    
    vendas_por_hora = db.session.query(
        db.func.extract('hour', Venda.data_hora_venda).label('hora'),
        db.func.count(Venda.id).label('quantidade')
    ).group_by('hora').all()
    
    return render_template('dashboard.html', 
                         total_vendas=total_vendas,
                         total_checkins=estatisticas['total_checkins'],
                         total_clientes=estatisticas['total_clientes'],
                         revendas_unicas=estatisticas['revendas_unicas'],
                         revendas_presentes=estatisticas['revendas_presentes'],
                         consultores_count=consultores_count,
                         vendas_por_revenda=vendas_por_revenda,
                         vendas_por_hora=vendas_por_hora,
                         now=agora())

@app.route('/importar-clientes', methods=['GET', 'POST'])
@login_required
@permissao_required('importacao')
def importar_clientes():
    if request.method == 'POST':
        if 'file' not in request.files:
            return 'Nenhum arquivo selecionado', 400
        
        file = request.files['file']
        if file.filename == '':
            return 'Nenhum arquivo selecionado', 400
        
        if file and file.filename.endswith('.xlsx'):
            try:
                data = read_excel_file(file)
                
                colunas_necessarias = ['CNPJ', 'RAZÃO SOCIAL', 'RESPONSÁVEL', 'CONSULTOR']
                if data and data[0]:
                    colunas_arquivo = [key.upper() for key in data[0].keys()]
                    if not all(coluna in colunas_arquivo for coluna in colunas_necessarias):
                        return f'Arquivo deve conter as colunas: {", ".join(colunas_necessarias)}', 400
                else:
                    return 'Arquivo vazio ou formato inválido', 400
                
                clientes_importados = 0
                clientes_atualizados = 0
                
                for row in data:
                    cnpj_normalizado = normalizar_cnpj(str(row.get('CNPJ', '')))
                    
                    cliente_existente = Cliente.query.filter_by(
                        cnpj=cnpj_normalizado,
                        responsavel=row.get('RESPONSÁVEL', '')
                    ).first()
                    
                    direito_imagem = False
                    if 'DIREITO IMAGEM' in row:
                        direito_imagem_val = str(row.get('DIREITO IMAGEM', '')).upper().strip()
                        direito_imagem = direito_imagem_val in ['SIM', 'YES', 'S', 'Y', '1', 'VERDADEIRO', 'TRUE']
                    
                    if cliente_existente:
                        cliente_existente.razao_social = row.get('RAZÃO SOCIAL', '')
                        cliente_existente.consultor = row.get('CONSULTOR', 'Não Informado')
                        cliente_existente.direito_imagem = direito_imagem
                        clientes_atualizados += 1
                    else:
                        cliente = Cliente(
                            cnpj=cnpj_normalizado,
                            razao_social=row.get('RAZÃO SOCIAL', ''),
                            responsavel=row.get('RESPONSÁVEL', ''),
                            consultor=row.get('CONSULTOR', 'Não Informado'),
                            direito_imagem=direito_imagem
                        )
                        db.session.add(cliente)
                        clientes_importados += 1
                
                db.session.commit()
                
                registrar_log('importacao_clientes', 'importacao', {
                    'clientes_importados': clientes_importados,
                    'clientes_atualizados': clientes_atualizados,
                    'total_registros': len(data)
                })
                
                mensagem = f'Importação concluída! '
                if clientes_importados > 0:
                    mensagem += f'{clientes_importados} novos clientes importados. '
                if clientes_atualizados > 0:
                    mensagem += f'{clientes_atualizados} clientes atualizados.'
                
                return mensagem
                
            except Exception as e:
                registrar_log('erro_importacao_clientes', 'importacao', {
                    'erro': str(e),
                    'arquivo': file.filename
                })
                return f'Erro na importação: {str(e)}', 500
    
    total_vendas = db.session.query(db.func.sum(Venda.valor_pedido)).scalar() or 0
    estatisticas = get_estatisticas_avancadas()
    consultores_count = db.session.query(Cliente.consultor).distinct().count()
    
    return render_template('import_clientes.html',
                         total_vendas=total_vendas,
                         total_checkins=estatisticas['total_checkins'],
                         total_clientes=estatisticas['total_clientes'],
                         revendas_unicas=estatisticas['revendas_unicas'],
                         revendas_presentes=estatisticas['revendas_presentes'],
                         consultores_count=consultores_count)

@app.route('/checkin', methods=['GET', 'POST'])
@login_required
@permissao_required('checkin')
def checkin():
    if request.method == 'POST':
        cnpj = request.form.get('cnpj')
        cnpj_limpo = ''.join(filter(str.isdigit, cnpj))
        
        clientes = Cliente.query.filter(
            db.or_(
                Cliente.cnpj == cnpj_limpo,
                Cliente.cnpj == cnpj
            )
        ).all()
        
        if not clientes:
            return jsonify({'success': False, 'message': 'CNPJ não encontrado na lista'})
        
        if len(clientes) == 1:
            cliente = clientes[0]
            
            if not cliente.direito_imagem:
                return jsonify({
                    'success': False, 
                    'need_direito_imagem': True,
                    'message': 'Cliente não possui termo de direito de imagem assinado',
                    'cliente_id': cliente.id,
                    'razao_social': cliente.razao_social,
                    'responsavel': cliente.responsavel
                })
            
            if not cliente.checkin_realizado:
                veio_carro = request.form.get('veio_carro') == 'true'
                placa_veiculo = request.form.get('placa_veiculo', '').upper().strip() if veio_carro else None
                
                cliente.checkin_realizado = True
                cliente.horario_checkin = agora()
                cliente.responsavel_checkin = cliente.responsavel
                cliente.veio_carro = veio_carro
                cliente.placa_veiculo = placa_veiculo
                
                db.session.commit()
                
                registrar_log('checkin_realizado', 'checkin', {
                    'cliente_id': cliente.id,
                    'cnpj': cliente.cnpj,
                    'razao_social': cliente.razao_social,
                    'responsavel': cliente.responsavel,
                    'veio_carro': veio_carro,
                    'placa_veiculo': placa_veiculo
                })
                
                return jsonify({
                    'success': True, 
                    'message': f'Check-in realizado para {cliente.responsavel}!',
                    'single': True
                })
            else:
                return jsonify({
                    'success': False, 
                    'message': f'Check-in já realizado para {cliente.responsavel}'
                })
        else:
            checkins_realizados = [c for c in clientes if c.checkin_realizado]
            
            return jsonify({
                'success': False,
                'multiple': True,
                'message': 'Encontramos múltiplos responsáveis para este CNPJ. Por favor, selecione:',
                'responsaveis': [{
                    'id': c.id,
                    'nome': c.responsavel,
                    'ja_checkin': c.checkin_realizado,
                    'direito_imagem': c.direito_imagem
                } for c in clientes]
            })
    
    total_clientes = Cliente.query.count()
    total_checkins = Cliente.query.filter_by(checkin_realizado=True).count()
    
    return render_template('checkin.html',
                         total_clientes=total_clientes,
                         total_checkins=total_checkins)
    
@app.route('/confirmar-direito-imagem', methods=['POST'])
@login_required
@permissao_required('checkin')
def confirmar_direito_imagem():
    cliente_id = request.form.get('cliente_id')
    
    cliente = Cliente.query.get(cliente_id)
    if not cliente:
        return jsonify({'success': False, 'message': 'Cliente não encontrado'})
    
    cliente.direito_imagem = True
    db.session.commit()
    
    registrar_log('direito_imagem_confirmado', 'checkin', {
        'cliente_id': cliente.id,
        'cnpj': cliente.cnpj,
        'razao_social': cliente.razao_social,
        'responsavel': cliente.responsavel
    })
    
    return jsonify({
        'success': True, 
        'message': 'Termo de direito de imagem confirmado!'
    })

@app.route('/api/todos-clientes')
@login_required
def api_todos_clientes():
    clientes = Cliente.query.order_by(Cliente.razao_social).all()
    
    result = []
    for cliente in clientes:
        result.append({
            'id': cliente.id,
            'cnpj': cliente.cnpj,
            'razao_social': cliente.razao_social,
            'responsavel': cliente.responsavel,
            'consultor': cliente.consultor,
            'checkin_realizado': cliente.checkin_realizado,
            'horario_checkin': cliente.horario_checkin.isoformat() if cliente.horario_checkin else None,
            'direito_imagem': cliente.direito_imagem,
            'veio_carro': cliente.veio_carro,
            'placa_veiculo': cliente.placa_veiculo
        })
    
    return jsonify(result)

@app.route('/checkin-responsavel', methods=['POST'])
@login_required
def checkin_responsavel():
    cliente_id = request.form.get('cliente_id')
    veio_carro = request.form.get('veio_carro') == 'true'
    placa_veiculo = request.form.get('placa_veiculo', '').upper().strip() if veio_carro else None
    
    cliente = Cliente.query.get(cliente_id)
    if cliente:
        if not cliente.checkin_realizado:
            if not cliente.direito_imagem:
                return jsonify({
                    'success': False, 
                    'message': 'Cliente não possui termo de direito de imagem assinado'
                })
            
            cliente.checkin_realizado = True
            cliente.horario_checkin = agora()
            cliente.responsavel_checkin = cliente.responsavel
            cliente.veio_carro = veio_carro
            cliente.placa_veiculo = placa_veiculo
            db.session.commit()
            
            registrar_log('checkin_responsavel', 'checkin', {
                'cliente_id': cliente.id,
                'cnpj': cliente.cnpj,
                'razao_social': cliente.razao_social,
                'responsavel': cliente.responsavel,
                'veio_carro': veio_carro,
                'placa_veiculo': placa_veiculo
            })
            
            return jsonify({
                'success': True, 
                'message': f'Check-in realizado para {cliente.responsavel}!'
            })
        else:
            return jsonify({
                'success': False, 
                'message': 'Check-in já realizado para este responsável'
            })
    
    return jsonify({'success': False, 'message': 'Responsável não encontrado'})

@app.route('/cadastro-rapido-checkin', methods=['POST'])
@login_required
def cadastro_rapido_checkin():
    data = request.get_json()
    
    cnpj = data.get('cnpj')
    razao_social = data.get('razao_social')
    responsavel = data.get('responsavel')
    consultor = data.get('consultor', 'Não Informado')
    direito_imagem = data.get('direito_imagem', False)
    
    if not cnpj or not razao_social or not responsavel:
        return jsonify({'success': False, 'message': 'CNPJ, Razão Social e Responsável são obrigatórios'})
    
    if not direito_imagem:
        return jsonify({'success': False, 'message': 'É necessário confirmar o termo de direito de imagem'})
    
    cnpj_normalizado = normalizar_cnpj(cnpj)
    
    try:
        cliente_existente = Cliente.query.filter_by(
            cnpj=cnpj_normalizado,
            responsavel=responsavel
        ).first()
        
        if cliente_existente:
            if not cliente_existente.checkin_realizado:
                cliente_existente.checkin_realizado = True
                cliente_existente.horario_checkin = agora()
                cliente_existente.responsavel_checkin = responsavel
                cliente_existente.direito_imagem = True
                db.session.commit()
                
                registrar_log('cadastro_rapido_checkin', 'checkin', {
                    'tipo': 'cliente_existente',
                    'cliente_id': cliente_existente.id,
                    'cnpj': cnpj_normalizado,
                    'razao_social': razao_social,
                    'responsavel': responsavel
                })
                
                return jsonify({
                    'success': True, 
                    'message': f'Check-in realizado para {responsavel} da revenda {razao_social}!'
                })
            else:
                return jsonify({
                    'success': False, 
                    'message': f'Check-in já realizado para {responsavel}'
                })
        else:
            revenda_existente = Cliente.query.filter_by(cnpj=cnpj_normalizado).first()
            
            if revenda_existente:
                novo_cliente = Cliente(
                    cnpj=cnpj_normalizado,
                    razao_social=razao_social,
                    responsavel=responsavel,
                    consultor=consultor,
                    checkin_realizado=True,
                    horario_checkin=agora(),
                    responsavel_checkin=responsavel,
                    direito_imagem=True
                )
                db.session.add(novo_cliente)
                db.session.commit()
                
                registrar_log('cadastro_rapido_checkin', 'checkin', {
                    'tipo': 'novo_responsavel',
                    'cliente_id': novo_cliente.id,
                    'cnpj': cnpj_normalizado,
                    'razao_social': razao_social,
                    'responsavel': responsavel
                })
                
                return jsonify({
                    'success': True, 
                    'message': f'Novo responsável {responsavel} cadastrado e check-in realizado para {razao_social}!'
                })
            else:
                novo_cliente = Cliente(
                    cnpj=cnpj_normalizado,
                    razao_social=razao_social,
                    responsavel=responsavel,
                    consultor=consultor,
                    checkin_realizado=True,
                    horario_checkin=agora(),
                    responsavel_checkin=responsavel,
                    direito_imagem=True
                )
                db.session.add(novo_cliente)
                db.session.commit()
                
                registrar_log('cadastro_rapido_checkin', 'checkin', {
                    'tipo': 'nova_revenda',
                    'cliente_id': novo_cliente.id,
                    'cnpj': cnpj_normalizado,
                    'razao_social': razao_social,
                    'responsavel': responsavel
                })
                
                return jsonify({
                    'success': True, 
                    'message': f'Nova revenda {razao_social} cadastrada e check-in realizado para {responsavel}!'
                })
                
    except Exception as e:
        registrar_log('erro_cadastro_rapido', 'checkin', {
            'cnpj': cnpj,
            'razao_social': razao_social,
            'responsavel': responsavel,
            'erro': str(e)
        })
        return jsonify({'success': False, 'message': f'Erro no cadastro: {str(e)}'})

@app.route('/registrar-venda', methods=['GET', 'POST'])
@login_required
@permissao_required('vendas')
def registrar_venda():
    if request.method == 'POST':
        cnpj_compra = request.form.get('cnpj_compra')
        numero_pedido = request.form.get('numero_pedido')
        valor_pedido = request.form.get('valor_pedido')
        produtos_data = request.form.get('produtos_data')
        
        if not numero_pedido.isdigit():
            return jsonify({'success': False, 'message': 'Número do pedido deve conter apenas números'})
        
        try:
            valor_pedido_float = float(valor_pedido)
            produtos = json.loads(produtos_data) if produtos_data else []
        except (ValueError, json.JSONDecodeError):
            return jsonify({'success': False, 'message': 'Dados inválidos'})
        
        pedido_existente = Venda.query.filter_by(numero_pedido=numero_pedido).first()
        if pedido_existente:
            return jsonify({'success': False, 'message': 'Número de pedido já registrado'})
        
        cliente = Cliente.query.filter_by(cnpj=cnpj_compra, checkin_realizado=True).first()
        
        if cliente:
            venda = Venda(
                cnpj_compra=cnpj_compra,
                numero_pedido=numero_pedido,
                valor_pedido=valor_pedido_float,
                cliente_id=cliente.id
            )
            db.session.add(venda)
            db.session.flush()
            
            produtos_vendidos = []
            for produto in produtos:
                equipamento_id = produto['equipamento_id']
                quantidade = produto['quantidade']
                
                if equipamento_id and quantidade:
                    equipamento = Estoque.query.get(equipamento_id)
                    if equipamento and equipamento.quantidade_disponivel >= quantidade:
                        venda_equipamento = VendaEquipamento(
                            venda_id=venda.id,
                            equipamento_id=equipamento_id,
                            quantidade=quantidade
                        )
                        db.session.add(venda_equipamento)
                        equipamento.quantidade_disponivel -= quantidade
                        
                        produtos_vendidos.append({
                            'equipamento': f"{equipamento.fabricante} - {equipamento.modelo}",
                            'quantidade': quantidade
                        })
                    else:
                        db.session.rollback()
                        return jsonify({'success': False, 'message': f'Estoque insuficiente para {equipamento.fabricante} - {equipamento.modelo}'})
            
            db.session.commit()
            
            registrar_log('venda_registrada', 'vendas', {
                'venda_id': venda.id,
                'numero_pedido': numero_pedido,
                'cnpj_compra': cnpj_compra,
                'valor_pedido': valor_pedido_float,
                'cliente_id': cliente.id,
                'cliente_razao_social': cliente.razao_social,
                'produtos': produtos_vendidos
            })
            
            mensagem = 'Venda registrada com sucesso!'
            if produtos:
                mensagem += ' Produtos debitados do estoque.'
            return jsonify({'success': True, 'message': mensagem})
        else:
            return jsonify({
                'success': False, 
                'need_link': True,
                'message': 'CNPJ não fez check-in. É necessário vincular com um CNPJ que fez check-in.'
            })
    
    return render_template('vendas.html')

@app.route('/vincular-cnpj', methods=['POST'])
@login_required
def vincular_cnpj():
    cnpj_checkin = request.form.get('cnpj_checkin')
    cnpj_compra = request.form.get('cnpj_compra')
    numero_pedido = request.form.get('numero_pedido')
    valor_pedido = request.form.get('valor_pedido')
    produtos_data = request.form.get('produtos_data')
    
    if not numero_pedido.isdigit():
        return jsonify({'success': False, 'message': 'Número do pedido deve conter apenas números'})
    
    try:
        produtos = json.loads(produtos_data) if produtos_data else []
    except json.JSONDecodeError:
        return jsonify({'success': False, 'message': 'Dados de produtos inválidos'})
    
    pedido_existente = Venda.query.filter_by(numero_pedido=numero_pedido).first()
    if pedido_existente:
        return jsonify({'success': False, 'message': 'Número de pedido já registrado'})
    
    cliente = Cliente.query.filter_by(cnpj=cnpj_checkin, checkin_realizado=True).first()
    
    if cliente:
        venda = Venda(
            cnpj_compra=cnpj_compra,
            numero_pedido=numero_pedido,
            valor_pedido=float(valor_pedido),
            cliente_id=cliente.id,
            cnpj_checkin_vinculado=cnpj_checkin
        )
        db.session.add(venda)
        db.session.flush()
        
        produtos_vendidos = []
        for produto in produtos:
            equipamento_id = produto['equipamento_id']
            quantidade = produto['quantidade']
            
            if equipamento_id and quantidade:
                equipamento = Estoque.query.get(equipamento_id)
                if equipamento and equipamento.quantidade_disponivel >= quantidade:
                    venda_equipamento = VendaEquipamento(
                        venda_id=venda.id,
                        equipamento_id=equipamento_id,
                        quantidade=quantidade
                    )
                    db.session.add(venda_equipamento)
                    equipamento.quantidade_disponivel -= quantidade
                    
                    produtos_vendidos.append({
                        'equipamento': f"{equipamento.fabricante} - {equipamento.modelo}",
                        'quantidade': quantidade
                    })
                else:
                    db.session.rollback()
                    return jsonify({'success': False, 'message': f'Estoque insuficiente para {equipamento.fabricante} - {equipamento.modelo}'})
        
        db.session.commit()
        
        registrar_log('venda_vinculada', 'vendas', {
            'venda_id': venda.id,
            'numero_pedido': numero_pedido,
            'cnpj_compra': cnpj_compra,
            'cnpj_checkin': cnpj_checkin,
            'valor_pedido': float(valor_pedido),
            'cliente_id': cliente.id,
            'cliente_razao_social': cliente.razao_social,
            'produtos': produtos_vendidos
        })
        
        return jsonify({'success': True, 'message': 'Venda vinculada e registrada com sucesso!'})
    else:
        return jsonify({'success': False, 'message': 'CNPJ de check-in não encontrado'})
    
@app.route('/api/verificar-pedido/<numero_pedido>')
@login_required
def api_verificar_pedido(numero_pedido):
    pedido_existente = Venda.query.filter_by(numero_pedido=numero_pedido).first()
    return jsonify({'existe': pedido_existente is not None})
    
@app.route('/estoque')
def estoque():
    equipamentos = Estoque.query.filter_by(ativo=True).order_by(Estoque.fabricante, Estoque.modelo).all()
    
    total_equipamentos = len(equipamentos)
    total_disponivel = sum(e.quantidade_disponivel for e in equipamentos)
    total_estoque = sum(e.quantidade_total for e in equipamentos)
    percentual_disponivel = (total_disponivel / total_estoque * 100) if total_estoque > 0 else 0
    
    return render_template('estoque.html',
                         equipamentos=equipamentos,
                         total_equipamentos=total_equipamentos,
                         total_disponivel=total_disponivel,
                         total_estoque=total_estoque,
                         percentual_disponivel=percentual_disponivel)

@app.route('/adicionar-equipamento', methods=['POST'])
@login_required
def adicionar_equipamento():
    fabricante = request.form.get('fabricante')
    modelo = request.form.get('modelo')
    quantidade = request.form.get('quantidade')
    
    if not fabricante or not modelo or not quantidade:
        return jsonify({'success': False, 'message': 'Preencha todos os campos obrigatórios'})
    
    try:
        quantidade_int = int(quantidade)
        
        equipamento_existente = Estoque.query.filter_by(
            fabricante=fabricante.upper(),
            modelo=modelo.upper(),
            ativo=True
        ).first()
        
        if equipamento_existente:
            equipamento_existente.quantidade_total += quantidade_int
            equipamento_existente.quantidade_disponivel += quantidade_int
            
            registrar_log('estoque_atualizado', 'estoque', {
                'equipamento_id': equipamento_existente.id,
                'fabricante': fabricante,
                'modelo': modelo,
                'quantidade_adicionada': quantidade_int,
                'novo_total': equipamento_existente.quantidade_total,
                'novo_disponivel': equipamento_existente.quantidade_disponivel,
                'tipo': 'atualizacao'
            })
        else:
            equipamento = Estoque(
                fabricante=fabricante.upper(),
                modelo=modelo.upper(),
                quantidade_total=quantidade_int,
                quantidade_disponivel=quantidade_int
            )
            db.session.add(equipamento)
            db.session.flush()
            
            registrar_log('equipamento_adicionado', 'estoque', {
                'equipamento_id': equipamento.id,
                'fabricante': fabricante,
                'modelo': modelo,
                'quantidade': quantidade_int,
                'tipo': 'novo'
            })
        
        db.session.commit()
        return jsonify({'success': True, 'message': 'Equipamento adicionado ao estoque com sucesso!'})
        
    except Exception as e:
        registrar_log('erro_adicionar_equipamento', 'estoque', {
            'fabricante': fabricante,
            'modelo': modelo,
            'quantidade': quantidade,
            'erro': str(e)
        })
        return jsonify({'success': False, 'message': f'Erro ao adicionar equipamento: {str(e)}'})

@app.route('/remover-equipamento/<int:equipamento_id>')
@login_required
def remover_equipamento(equipamento_id):
    equipamento = Estoque.query.get(equipamento_id)
    if equipamento:
        equipamento.ativo = False
        db.session.commit()
        
        registrar_log('equipamento_removido', 'estoque', {
            'equipamento_id': equipamento_id,
            'fabricante': equipamento.fabricante,
            'modelo': equipamento.modelo,
            'quantidade_total': equipamento.quantidade_total,
            'quantidade_disponivel': equipamento.quantidade_disponivel
        })
        
        return jsonify({'success': True, 'message': 'Equipamento removido do estoque!'})
    return jsonify({'success': False, 'message': 'Equipamento não encontrado'})

@app.route('/api/estoque-atual')
def api_estoque_atual():
    equipamentos = Estoque.query.filter_by(ativo=True).order_by(Estoque.fabricante, Estoque.modelo).all()
    
    result = []
    for equipamento in equipamentos:
        result.append({
            'id': equipamento.id,
            'fabricante': equipamento.fabricante,
            'modelo': equipamento.modelo,
            'quantidade_total': equipamento.quantidade_total,
            'quantidade_disponivel': equipamento.quantidade_disponivel,
            'status': 'success' if equipamento.quantidade_disponivel > 5 else 
                     'warning' if equipamento.quantidade_disponivel > 0 else 
                     'danger'
        })
    
    return jsonify(result)

@app.route('/estoque-publico')
def estoque_publico():
    return render_template('estoque_publico.html')

@app.route('/api/vendas-hoje')
@login_required
def api_vendas_hoje():
    hoje = agora().date()
    quantidade_vendas_hoje = Venda.query.filter(
        db.cast(Venda.data_hora_venda, db.Date) == hoje
    ).count()
    return jsonify({'total': quantidade_vendas_hoje})

@app.route('/api/brindes-sorteados')
@login_required
def api_brindes_sorteados():
    total_sorteados = Sorteio.query.count()
    return jsonify({'total': total_sorteados})

@app.route('/api/ticket-medio')
@login_required
def api_ticket_medio():
    total_vendas = db.session.query(db.func.sum(Venda.valor_pedido)).scalar() or 0
    count_vendas = Venda.query.count()
    media = total_vendas / count_vendas if count_vendas > 0 else 0
    return jsonify({'media': float(media)})

@app.route('/api/total-pedidos')
@login_required
def api_total_pedidos():
    total_pedidos = Venda.query.count()
    return jsonify({'total': total_pedidos})

@app.route('/sorteio')
@login_required
@permissao_required('sorteio')
def sorteio():
    atualizar_faturamento_sorteio()
    revendas_20k = get_revendas_para_sorteio('20k')
    revendas_50k = get_revendas_para_sorteio('50k')
    brindes_20k = Brinde.query.filter_by(tipo_sorteio='20k', ativo=True).all()
    brindes_50k = Brinde.query.filter_by(tipo_sorteio='50k', ativo=True).all()
    
    total_participacoes_20k = len(revendas_20k)
    total_participacoes_50k = sum(revenda['participacoes'] for revenda in revendas_50k)
    
    brindes_20k_disponiveis = sum(1 for b in brindes_20k if b.quantidade_disponivel > 0)
    brindes_50k_disponiveis = sum(1 for b in brindes_50k if b.quantidade_disponivel > 0)
    
    ultimos_sorteios = Sorteio.query.order_by(Sorteio.data_sorteio.desc()).limit(10).all()
    
    return render_template('sorteio.html', 
                         revendas_20k=revendas_20k,
                         revendas_50k=revendas_50k,
                         brindes_20k=brindes_20k,
                         brindes_50k=brindes_50k,
                         total_revendas_20k=len(revendas_20k),
                         total_revendas_50k=len(revendas_50k),
                         total_participacoes_20k=total_participacoes_20k,
                         total_participacoes_50k=total_participacoes_50k,
                         brindes_20k_disponiveis=brindes_20k_disponiveis,
                         brindes_50k_disponiveis=brindes_50k_disponiveis,
                         ultimos_sorteios=ultimos_sorteios)

@app.route('/realizar-sorteio', methods=['POST'])
@login_required
def realizar_sorteio():
    tipo_brinde = request.form.get('tipo_brinde')
    cnpj_vencedor = request.form.get('cnpj_vencedor')
    responsavel = request.form.get('responsavel')
    
    if not cnpj_vencedor:
        return jsonify({'success': False, 'message': 'Por favor, selecione uma revenda para o sorteio'})
    
    if not responsavel:
        return jsonify({'success': False, 'message': 'Por favor, informe o responsável pelo recebimento'})
    
    revendas_qualificadas = get_revendas_para_sorteio(tipo_brinde)
    revenda_vencedora = next((r for r in revendas_qualificadas if r['cnpj'] == cnpj_vencedor), None)
    
    if not revenda_vencedora:
        return jsonify({'success': False, 'message': 'Revenda não encontrada ou não qualificada para este sorteio'})
    
    brindes_disponiveis = Brinde.query.filter_by(
        tipo_sorteio=tipo_brinde, 
        ativo=True
    ).filter(Brinde.quantidade_disponivel > 0).all()
    
    if not brindes_disponiveis:
        return jsonify({'success': False, 'message': f'Nenhum brinde disponível com estoque para a categoria R$ {tipo_brinde}'})
    
    brinde_sorteado = random.choice(brindes_disponiveis)
    brinde_sorteado.quantidade_disponivel -= 1
    
    sorteio = Sorteio(
        tipo_brinde=tipo_brinde,
        cnpj_vencedor=revenda_vencedora['cnpj'],
        razao_social_vencedor=revenda_vencedora['razao_social'],
        responsavel_recebimento=responsavel,
        valor_acumulado_revenda=revenda_vencedora['faturamento_total'],
        brinde_id=brinde_sorteado.id
    )
    db.session.add(sorteio)
    db.session.commit()
    
    pedidos_qualificadores = Venda.query.join(Cliente)\
                                       .filter(Cliente.cnpj == cnpj_vencedor)\
                                       .order_by(Venda.data_hora_venda.asc())\
                                       .all()
    
    participacoes_restantes = 0
    if tipo_brinde == '50k':
        sorteios_realizados = Sorteio.query.filter_by(
            cnpj_vencedor=cnpj_vencedor, 
            tipo_brinde='50k'
        ).count()
        
        participacoes_totais = revenda_vencedora.get('participacoes', 0)
        participacoes_restantes = max(0, participacoes_totais - sorteios_realizados)
    
    registrar_log('sorteio_realizado', 'sorteio', {
        'sorteio_id': sorteio.id,
        'tipo_brinde': tipo_brinde,
        'cnpj_vencedor': cnpj_vencedor,
        'razao_social_vencedor': revenda_vencedora['razao_social'],
        'responsavel_recebimento': responsavel,
        'valor_acumulado': revenda_vencedora['faturamento_total'],
        'participacoes_antes': revenda_vencedora.get('participacoes', 1),
        'participacoes_restantes': participacoes_restantes,
        'brinde_nome': brinde_sorteado.nome,
        'brinde_valor': brinde_sorteado.valor_aproximado,
        'quantidade_pedidos_qualificadores': len(pedidos_qualificadores)
    })
    
    return jsonify({
        'success': True, 
        'message': f'Sorteio de R$ {tipo_brinde} realizado!' + 
                  (f' (Participações restantes: {participacoes_restantes})' if tipo_brinde == '50k' and participacoes_restantes > 0 else ''),
        'vencedor': {
            'razao_social': revenda_vencedora['razao_social'],
            'cnpj': revenda_vencedora['cnpj'],
            'faturamento': revenda_vencedora['faturamento_total'],
            'participacoes': revenda_vencedora.get('participacoes', 1)
        },
        'brinde': {
            'nome': brinde_sorteado.nome,
            'descricao': brinde_sorteado.descricao,
            'valor_aproximado': brinde_sorteado.valor_aproximado
        },
        'participacoes_restantes': participacoes_restantes,
        'pedidos_qualificadores': [{
            'numero_pedido': p.numero_pedido,
            'valor': p.valor_pedido,
            'data': p.data_hora_venda.strftime('%d/%m/%Y %H:%M')
        } for p in pedidos_qualificadores]
    })

@app.route('/confirmar-entrega')
@login_required
@permissao_required('entrega_brindes')
def confirmar_entrega_page():
    return render_template('confirmar_entrega.html')

@app.route('/api/brindes-sorteados-completo')
@login_required
def api_brindes_sorteados_completo():
    sorteios = Sorteio.query.order_by(Sorteio.data_sorteio.desc()).all()
    
    result = []
    for sorteio in sorteios:
        result.append({
            'id': sorteio.id,
            'tipo_brinde': sorteio.tipo_brinde,
            'cnpj_vencedor': sorteio.cnpj_vencedor,
            'razao_social_vencedor': sorteio.razao_social_vencedor,
            'responsavel_recebimento': sorteio.responsavel_recebimento,
            'data_sorteio': sorteio.data_sorteio.isoformat() if sorteio.data_sorteio else None,
            'valor_acumulado_revenda': sorteio.valor_acumulado_revenda,
            'brinde_id': sorteio.brinde_id,
            'brinde_nome': sorteio.brinde.nome if sorteio.brinde else None,
            'brinde_descricao': sorteio.brinde.descricao if sorteio.brinde else None,
            'brinde_valor': sorteio.brinde.valor_aproximado if sorteio.brinde else None,
            'entregue': sorteio.entregue,
            'data_entrega': sorteio.data_entrega.isoformat() if sorteio.data_entrega else None,
            'responsavel_entrega': sorteio.responsavel_entrega,
            'observacao_entrega': sorteio.observacao_entrega
        })
    
    return jsonify(result)

@app.route('/confirmar-entrega', methods=['POST'])
@login_required
def confirmar_entrega():
    sorteio_id = request.form.get('sorteio_id')
    responsavel_entrega = request.form.get('responsavel_entrega')
    observacao_entrega = request.form.get('observacao_entrega')
    
    if not sorteio_id or not responsavel_entrega:
        return jsonify({'success': False, 'message': 'ID do sorteio e responsável pela entrega são obrigatórios'})
    
    try:
        sorteio = Sorteio.query.get(int(sorteio_id))
        if not sorteio:
            return jsonify({'success': False, 'message': 'Sorteio não encontrado'})
        
        if sorteio.entregue:
            return jsonify({'success': False, 'message': 'Este brinde já foi entregue'})
        
        sorteio.entregue = True
        sorteio.data_entrega = agora()
        sorteio.responsavel_entrega = responsavel_entrega
        sorteio.observacao_entrega = observacao_entrega
        
        db.session.commit()
        
        registrar_log('entrega_brinde_confirmada', 'entrega_brindes', {
            'sorteio_id': sorteio_id,
            'cnpj_vencedor': sorteio.cnpj_vencedor,
            'razao_social_vencedor': sorteio.razao_social_vencedor,
            'brinde_nome': sorteio.brinde.nome if sorteio.brinde else 'N/A',
            'responsavel_entrega': responsavel_entrega,
            'observacao': observacao_entrega
        })
        
        return jsonify({'success': True, 'message': 'Entrega confirmada com sucesso!'})
        
    except Exception as e:
        registrar_log('erro_confirmar_entrega', 'entrega_brindes', {
            'sorteio_id': sorteio_id,
            'erro': str(e)
        })
        return jsonify({'success': False, 'message': f'Erro ao confirmar entrega: {str(e)}'})

@app.route('/brindes')
@login_required
@permissao_required('brindes')
def brindes():
    brindes_20k = Brinde.query.filter_by(tipo_sorteio='20k', ativo=True).all()
    brindes_50k = Brinde.query.filter_by(tipo_sorteio='50k', ativo=True).all()
    return render_template('brindes.html', 
                         brindes_20k=brindes_20k, 
                         brindes_50k=brindes_50k)

@app.route('/adicionar-brinde', methods=['POST'])
@login_required
def adicionar_brinde():
    tipo_sorteio = request.form.get('tipo_sorteio')
    nome = request.form.get('nome')
    descricao = request.form.get('descricao')
    valor_aproximado = request.form.get('valor_aproximado')
    quantidade_total = request.form.get('quantidade_total')
    
    brinde = Brinde(
        tipo_sorteio=tipo_sorteio,
        nome=nome,
        descricao=descricao,
        valor_aproximado=float(valor_aproximado) if valor_aproximado else None,
        quantidade_total=int(quantidade_total),
        quantidade_disponivel=int(quantidade_total)
    )
    db.session.add(brinde)
    db.session.commit()
    
    registrar_log('brinde_adicionado', 'brindes', {
        'brinde_id': brinde.id,
        'tipo_sorteio': tipo_sorteio,
        'nome': nome,
        'valor_aproximado': float(valor_aproximado) if valor_aproximado else None,
        'quantidade_total': int(quantidade_total)
    })
    
    return jsonify({'success': True, 'message': 'Brinde adicionado com sucesso!'})

@app.route('/remover-brinde/<int:brinde_id>')
@login_required
def remover_brinde(brinde_id):
    brinde = Brinde.query.get(brinde_id)
    if brinde:
        brinde.ativo = False
        db.session.commit()
        
        registrar_log('brinde_removido', 'brindes', {
            'brinde_id': brinde_id,
            'nome': brinde.nome,
            'tipo_sorteio': brinde.tipo_sorteio
        })
        
        return jsonify({'success': True, 'message': 'Brinde removido com sucesso!'})
    return jsonify({'success': False, 'message': 'Brinde não encontrado'})

@app.route('/relatorios')
@login_required
@permissao_required('relatorios')
def relatorios():
    sorteios = Sorteio.query.order_by(Sorteio.data_sorteio.desc()).all()
    checkins = Cliente.query.filter_by(checkin_realizado=True)\
                          .order_by(Cliente.horario_checkin.desc())\
                          .all()
    vendas = Venda.query.join(Cliente)\
                       .order_by(Venda.data_hora_venda.desc())\
                       .all()
    
    sorteios_20k = [s for s in sorteios if s.tipo_brinde == '20k']
    sorteios_50k = [s for s in sorteios if s.tipo_brinde == '50k']
    total_faturamento_sorteios = sum(s.valor_acumulado_revenda for s in sorteios)
    
    revendas_unicas_checkin = db.session.query(Cliente.cnpj)\
                                       .filter(Cliente.checkin_realizado == True)\
                                       .distinct()\
                                       .count()
    responsaveis_unicos_checkin = db.session.query(Cliente.responsavel_checkin)\
                                           .filter(Cliente.checkin_realizado == True)\
                                           .distinct()\
                                           .count()
    total_clientes = Cliente.query.count()
    
    total_vendas_valor = sum(v.valor_pedido for v in vendas)
    revendas_compradoras = db.session.query(Venda.cnpj_compra).distinct().count()
    ticket_medio = total_vendas_valor / len(vendas) if vendas else 0
    
    return render_template('relatorios.html',
                         sorteios=sorteios,
                         sorteios_20k=sorteios_20k,
                         sorteios_50k=sorteios_50k,
                         total_faturamento_sorteios=total_faturamento_sorteios,
                         checkins=checkins,
                         revendas_unicas_checkin=revendas_unicas_checkin,
                         responsaveis_unicos_checkin=responsaveis_unicos_checkin,
                         total_clientes=total_clientes,
                         vendas=vendas,
                         total_vendas_valor=total_vendas_valor,
                         revendas_compradoras=revendas_compradoras,
                         ticket_medio=ticket_medio)

@app.route('/exportar-sorteios')
@login_required
@permissao_required('exportacao')
def exportar_sorteios():
    """Exporta TODOS os sorteios para Excel"""
    
    try:
        sorteios = Sorteio.query.options(db.joinedload(Sorteio.brinde)).order_by(Sorteio.data_sorteio.desc()).all()
        
        data = []
        for sorteio in sorteios:
            # Busca as vendas que qualificaram este sorteio
            vendas_qualificadoras = Venda.query.join(Cliente).filter(Cliente.cnpj == sorteio.cnpj_vencedor).all()
            
            data.append({
                'ID Sorteio': sorteio.id,
                'Tipo Brinde': f"R$ {sorteio.tipo_brinde}",
                'CNPJ Vencedor': sorteio.cnpj_vencedor,
                'Razão Social': sorteio.razao_social_vencedor,
                'Responsável Recebimento': sorteio.responsavel_recebimento,
                'Faturamento Acumulado': sorteio.valor_acumulado_revenda,
                'Brinde': sorteio.brinde.nome if sorteio.brinde else 'N/A',
                'Valor Brinde': sorteio.brinde.valor_aproximado if sorteio.brinde else 'N/A',
                'Data Sorteio': sorteio.data_sorteio.strftime('%d/%m/%Y %H:%M'),
                'Quantidade Pedidos': len(vendas_qualificadoras),
                'Números Pedidos': ', '.join([v.numero_pedido for v in vendas_qualificadoras]),
                'Valor Total Pedidos': sum(v.valor_pedido for v in vendas_qualificadoras),
                'Entregue': 'Sim' if sorteio.entregue else 'Não',
                'Data Entrega': sorteio.data_entrega.strftime('%d/%m/%Y %H:%M') if sorteio.data_entrega else 'N/A',
                'Responsável Entrega': sorteio.responsavel_entrega or 'N/A'
            })
        
        output = export_to_excel(data, 'todos_sorteios_ro_experience.xlsx', 'Sorteios')
        
        registrar_log('exportacao_todos_sorteios', 'exportacao', {
            'quantidade_sorteios': len(sorteios),
            'tipo_exportacao': 'excel'
        })
        
        return send_file(
            output,
            download_name='todos_sorteios_ro_experience.xlsx',
            as_attachment=True,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        print(f"❌ Erro ao exportar sorteios: {e}")
        flash(f'Erro ao exportar sorteios: {str(e)}', 'error')
        return redirect(url_for('sorteio'))


@app.route('/api/total-vendas')
@login_required
def api_total_vendas():
    total = db.session.query(db.func.sum(Venda.valor_pedido)).scalar() or 0
    return jsonify({'total_vendas': float(total)})

@app.route('/api/ultimos-checkins')
def api_ultimos_checkins():
    ultimos = Cliente.query.filter_by(checkin_realizado=True)\
                          .order_by(Cliente.horario_checkin.desc())\
                          .limit(10)\
                          .all()
    
    result = []
    for cliente in ultimos:
        result.append({
            'cnpj': cliente.cnpj,
            'razao_social': cliente.razao_social,
            'responsavel_checkin': cliente.responsavel_checkin or cliente.responsavel,
            'consultor': cliente.consultor,
            'horario_checkin': cliente.horario_checkin.isoformat() if cliente.horario_checkin else '',
            'veio_carro': cliente.veio_carro,
            'placa_veiculo': cliente.placa_veiculo
        })
    
    return jsonify(result)

@app.route('/api/buscar-veiculo/<placa>')
@login_required
def api_buscar_veiculo(placa):
    if not placa or len(placa.strip()) < 3:
        return jsonify({'success': False, 'message': 'Digite pelo menos 3 caracteres da placa'})
    
    placa_busca = placa.upper().strip()
    
    clientes = Cliente.query.filter(
        Cliente.checkin_realizado == True,
        Cliente.veio_carro == True,
        Cliente.placa_veiculo.ilike(f'%{placa_busca}%')
    ).all()
    
    if not clientes:
        return jsonify({'success': False, 'message': 'Nenhum veículo encontrado com esta placa'})
    
    result = []
    for cliente in clientes:
        result.append({
            'razao_social': cliente.razao_social,
            'responsavel': cliente.responsavel_checkin or cliente.responsavel,
            'cnpj': cliente.cnpj,
            'placa_veiculo': cliente.placa_veiculo,
            'horario_checkin': cliente.horario_checkin.strftime('%d/%m/%Y %H:%M') if cliente.horario_checkin else '',
            'consultor': cliente.consultor
        })
    
    return jsonify({'success': True, 'veiculos': result})

@app.route('/api/ultimos-clientes')
@login_required
def api_ultimos_clientes():
    ultimos = Cliente.query.order_by(Cliente.id.desc()).limit(15).all()
    
    result = []
    for cliente in ultimos:
        result.append({
            'cnpj': cliente.cnpj,
            'razao_social': cliente.razao_social,
            'responsavel': cliente.responsavel,
            'consultor': cliente.consultor,
            'checkin_realizado': cliente.checkin_realizado
        })
    
    return jsonify(result)

@app.route('/exportar-clientes')
@login_required
@permissao_required('exportacao')
def exportar_clientes():
    clientes = Cliente.query.all()
    
    data = []
    for cliente in clientes:
        data.append({
            'CNPJ': cliente.cnpj,
            'RAZÃO SOCIAL': cliente.razao_social,
            'RESPONSÁVEL': cliente.responsavel,
            'CONSULTOR': cliente.consultor,
            'Check-in Realizado': 'Sim' if cliente.checkin_realizado else 'Não',
            'Responsável Check-in': cliente.responsavel_checkin or '',
            'Horário Check-in': cliente.horario_checkin.strftime('%d/%m/%Y %H:%M') if cliente.horario_checkin else ''
        })
    
    output = export_to_excel(data, 'clientes_ro_experience.xlsx', 'Clientes')
    
    registrar_log('exportacao_clientes', 'exportacao', {
        'quantidade_clientes': len(clientes),
        'tipo_exportacao': 'clientes'
    })
    
    return send_file(output, 
                    download_name='clientes_ro_experience.xlsx',
                    as_attachment=True)

@app.route('/exportar-vendas')
@login_required
@permissao_required('exportacao')
def exportar_vendas():
    vendas = Venda.query.join(Cliente).all()
    
    data = []
    for venda in vendas:
        data.append({
            'CNPJ Compra': venda.cnpj_compra,
            'Número Pedido': venda.numero_pedido,
            'Valor Pedido': venda.valor_pedido,
            'Data/Hora': venda.data_hora_venda.strftime('%d/%m/%Y %H:%M'),
            'Razão Social': venda.cliente.razao_social,
            'Responsável': venda.cliente.responsavel,
            'Consultor': venda.cliente.consultor,
            'CNPJ Check-in Vinculado': venda.cnpj_checkin_vinculado or ''
        })
    
    output = export_to_excel(data, 'vendas_ro_experience.xlsx', 'Vendas')
    
    registrar_log('exportacao_vendas', 'exportacao', {
        'quantidade_vendas': len(vendas),
        'tipo_exportacao': 'vendas'
    })
    
    return send_file(output, 
                    download_name='vendas_ro_experience.xlsx',
                    as_attachment=True)

@app.route('/exportar-checkins')
@login_required
@permissao_required('exportacao')
def exportar_checkins():
    checkins = Cliente.query.filter_by(checkin_realizado=True).all()
    
    data = []
    for checkin in checkins:
        data.append({
            'CNPJ': checkin.cnpj,
            'Razão Social': checkin.razao_social,
            'Responsável Check-in': checkin.responsavel_checkin or checkin.responsavel,
            'Consultor': checkin.consultor,
            'Horário Check-in': checkin.horario_checkin.strftime('%d/%m/%Y %H:%M') if checkin.horario_checkin else ''
        })
    
    output = export_to_excel(data, 'checkins_ro_experience.xlsx', 'Check-ins')
    
    registrar_log('exportacao_checkins', 'exportacao', {
        'quantidade_checkins': len(checkins),
        'tipo_exportacao': 'checkins'
    })
    
    return send_file(output, 
                    download_name='checkins_ro_experience.xlsx',
                    as_attachment=True)

@app.route('/download-template')
@login_required
def download_template():
    data = [
        {
            'CNPJ': '12.345.678/0001-90',
            'RAZÃO SOCIAL': 'Empresa ABC Ltda',
            'RESPONSÁVEL': 'João Silva',
            'CONSULTOR': 'Consultor A',
            'DIREITO IMAGEM': 'SIM'
        },
        {
            'CNPJ': '12.345.678/0001-90',
            'RAZÃO SOCIAL': 'Empresa ABC Ltda',
            'RESPONSÁVEL': 'Maria Santos',
            'CONSULTOR': 'Consultor A',
            'DIREITO IMAGEM': 'NÃO'
        },
        {
            'CNPJ': '98.765.432/0001-10',
            'RAZÃO SOCIAL': 'Comércio DEF S/A',
            'RESPONSÁVEL': 'Pedro Oliveira',
            'CONSULTOR': 'Consultor B',
            'DIREITO IMAGEM': 'SIM'
        },
        {
            'CNPJ': '11.222.333/0001-44',
            'RAZÃO SOCIAL': 'Indústria GHI ME',
            'RESPONSÁVEL': 'Carlos Souza',
            'CONSULTOR': 'Consultor C',
            'DIREITO IMAGEM': 'NÃO'
        },
        {
            'CNPJ': '55.666.777/0001-88',
            'RAZÃO SOCIAL': 'Serviços JKL EIRELI',
            'RESPONSÁVEL': 'Ana Paula Santos',
            'CONSULTOR': 'Consultor A',
            'DIREITO IMAGEM': 'SIM'
        }
    ]
    
    output = export_to_excel(data, 'template_clientes_ro_experience.xlsx', 'Template_Clientes')
    
    registrar_log('template_download', 'exportacao', {
        'tipo_template': 'clientes_com_direito_imagem'
    })
    
    return send_file(
        output,
        download_name='template_clientes_ro_experience.xlsx',
        as_attachment=True,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

@app.route('/login', methods=['GET', 'POST'])
def login():
    if 'usuario_id' in session:
        return redirect(url_for('index'))
    
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        
        usuario = Usuario.query.filter_by(username=username, ativo=True).first()
        
        if usuario and check_password_hash(usuario.password_hash, password):
            session['usuario_id'] = usuario.id
            session['username'] = usuario.username
            session['nome'] = usuario.nome
            session['nivel_acesso'] = usuario.nivel_acesso
            
            registrar_log('login', 'autenticacao', {
                'usuario': usuario.username,
                'nivel_acesso': usuario.nivel_acesso
            })
            
            return redirect(url_for('index'))
        else:
            return render_template('login.html', error='Credenciais inválidas')
    
    return render_template('login.html')

@app.route('/logout')
def logout():
    if 'usuario_id' in session:
        registrar_log('logout', 'autenticacao', {
            'usuario': session.get('username')
        })
        
        session.clear()
    
    return redirect(url_for('login'))

@app.route('/alterar-senha', methods=['GET', 'POST'])
@login_required
def alterar_senha():
    if request.method == 'POST':
        senha_atual = request.form.get('senha_atual')
        nova_senha = request.form.get('nova_senha')
        confirmar_senha = request.form.get('confirmar_senha')
        
        usuario = Usuario.query.get(session['usuario_id'])
        
        if not check_password_hash(usuario.password_hash, senha_atual):
            return jsonify({'success': False, 'message': 'Senha atual incorreta'})
        
        if nova_senha != confirmar_senha:
            return jsonify({'success': False, 'message': 'As senhas não coincidem'})
        
        if len(nova_senha) < 6:
            return jsonify({'success': False, 'message': 'A senha deve ter pelo menos 6 caracteres'})
        
        usuario.password_hash = generate_password_hash(nova_senha)
        db.session.commit()
        
        registrar_log('alteracao_senha', 'autenticacao', {
            'usuario': usuario.username
        })
        
        return jsonify({'success': True, 'message': 'Senha alterada com sucesso!'})
    
    return render_template('alterar_senha.html')

@app.route('/gestao-usuarios')
@login_required
@admin_required
def gestao_usuarios():
    usuarios = Usuario.query.order_by(Usuario.data_criacao.desc()).all()
    return render_template('gestao_usuarios.html', 
                         usuarios=usuarios, 
                         modulos_sistema=MODULOS_SISTEMA)

@app.route('/criar-usuario', methods=['POST'])
@login_required
@admin_required
def criar_usuario():
    username = request.form.get('username')
    nome = request.form.get('nome')
    password = request.form.get('password')
    nivel_acesso = request.form.get('nivel_acesso', 'operador')
    permissoes_selecionadas = request.form.getlist('permissoes')
    
    if Usuario.query.filter_by(username=username).first():
        flash('Usuário já existe', 'error')
        return redirect(url_for('gestao_usuarios'))
    
    if len(password) < 6:
        flash('A senha deve ter pelo menos 6 caracteres', 'error')
        return redirect(url_for('gestao_usuarios'))
    
    permissoes = {}
    if nivel_acesso == 'admin':
        permissoes = {modulo: True for modulo in MODULOS_SISTEMA.keys()}
    else:
        permissoes = {modulo: (modulo in permissoes_selecionadas) for modulo in MODULOS_SISTEMA.keys()}
    
    usuario = Usuario(
        username=username,
        nome=nome,
        password_hash=generate_password_hash(password),
        nivel_acesso=nivel_acesso,
        permissoes=json.dumps(permissoes)
    )
    
    db.session.add(usuario)
    db.session.commit()
    
    registrar_log('criacao_usuario', 'usuarios', {
        'usuario_criado': username,
        'nivel_acesso': nivel_acesso,
        'permissoes': permissoes_selecionadas,
        'criado_por': session.get('username')
    })
    
    flash('Usuário criado com sucesso!', 'success')
    return redirect(url_for('gestao_usuarios'))

@app.route('/editar-usuario', methods=['POST'])
@login_required
@admin_required
def editar_usuario():
    usuario_id = request.form.get('usuario_id')
    nome = request.form.get('nome')
    nivel_acesso = request.form.get('nivel_acesso')
    password = request.form.get('password')
    permissoes_selecionadas = request.form.getlist('permissoes')
    
    usuario = Usuario.query.get(usuario_id)
    if not usuario:
        flash('Usuário não encontrado', 'error')
        return redirect(url_for('gestao_usuarios'))
    
    usuario.nome = nome
    usuario.nivel_acesso = nivel_acesso
    
    if nivel_acesso == 'admin':
        permissoes = {modulo: True for modulo in MODULOS_SISTEMA.keys()}
    else:
        permissoes = {modulo: (modulo in permissoes_selecionadas) for modulo in MODULOS_SISTEMA.keys()}
    
    usuario.permissoes = json.dumps(permissoes)
    
    if password and len(password) >= 6:
        usuario.password_hash = generate_password_hash(password)
    
    db.session.commit()
    
    registrar_log('edicao_usuario', 'usuarios', {
        'usuario_editado': usuario.username,
        'nivel_acesso': nivel_acesso,
        'permissoes': permissoes_selecionadas,
        'editado_por': session.get('username')
    })
    
    flash('Usuário atualizado com sucesso!', 'success')
    return redirect(url_for('gestao_usuarios'))

@app.route('/alternar-status-usuario/<int:usuario_id>', methods=['POST'])
@login_required
@admin_required
def alternar_status_usuario(usuario_id):
    usuario = Usuario.query.get(usuario_id)
    if not usuario:
        return jsonify({'success': False, 'message': 'Usuário não encontrado'})
    
    if usuario.id == session['usuario_id']:
        return jsonify({'success': False, 'message': 'Não é possível inativar seu próprio usuário'})
    
    usuario.ativo = not usuario.ativo
    db.session.commit()
    
    acao = 'ativado' if usuario.ativo else 'inativado'
    
    registrar_log(f'usuario_{acao}', 'usuarios', {
        'usuario': usuario.username,
        'acao': acao,
        'realizado_por': session.get('username')
    })
    
    return jsonify({'success': True, 'message': f'Usuário {acao} com sucesso!'})

@app.route('/api/usuario/<int:usuario_id>')
@login_required
@admin_required
def api_usuario(usuario_id):
    usuario = Usuario.query.get(usuario_id)
    if usuario:
        return jsonify({
            'id': usuario.id,
            'username': usuario.username,
            'nome': usuario.nome,
            'nivel_acesso': usuario.nivel_acesso,
            'ativo': usuario.ativo
        })
    return jsonify({'error': 'Usuário não encontrado'}), 404

@app.route('/excluir-usuario/<int:usuario_id>', methods=['POST'])
@login_required
@admin_required
def excluir_usuario(usuario_id):
    usuario = Usuario.query.get(usuario_id)
    if not usuario:
        return jsonify({'success': False, 'message': 'Usuário não encontrado'})
    
    if usuario.id == session['usuario_id']:
        return jsonify({'success': False, 'message': 'Não é possível excluir seu próprio usuário'})
    
    if usuario.nivel_acesso == 'admin':
        total_admins = Usuario.query.filter_by(nivel_acesso='admin', ativo=True).count()
        if total_admins <= 1:
            return jsonify({'success': False, 'message': 'Não é possível excluir o último administrador do sistema'})
    
    registrar_log('exclusao_usuario', 'usuarios', {
        'usuario_excluido': usuario.username,
        'nome': usuario.nome,
        'nivel_acesso': usuario.nivel_acesso,
        'excluido_por': session.get('username')
    })
    
    db.session.delete(usuario)
    db.session.commit()
    
    return jsonify({'success': True, 'message': 'Usuário excluído com sucesso!'})

@app.route('/api/log-detalhes/<int:log_id>')
@login_required
@admin_required
def api_log_detalhes(log_id):
    log = LogAuditoria.query.get(log_id)
    if log:
        return jsonify({
            'id': log.id,
            'dados': log.dados
        })
    return jsonify({'error': 'Log não encontrado'}), 404

@app.context_processor
def utility_processor():
    def check_permission_template(modulo):
        return tem_permissao(modulo)
    return dict(tem_permissao=check_permission_template)

@app.route('/api/metricas-ultima-hora')
@login_required
def api_metricas_ultima_hora():
    uma_hora_atras = agora() - timedelta(hours=1)
    
    checkins_ultima_hora = Cliente.query.filter(
        Cliente.horario_checkin >= uma_hora_atras
    ).count()
    
    vendas_ultima_hora = Venda.query.filter(
        Venda.data_hora_venda >= uma_hora_atras
    ).count()
    
    sorteios_ultima_hora = Sorteio.query.filter(
        Sorteio.data_sorteio >= uma_hora_atras
    ).count()
    
    return jsonify({
        'checkins': checkins_ultima_hora,
        'vendas': vendas_ultima_hora,
        'sorteios': sorteios_ultima_hora
    })

@app.route('/logs-auditoria')
@login_required
@admin_required
def logs_auditoria():
    logs = LogAuditoria.query.order_by(LogAuditoria.data_hora.desc()).limit(100).all()
    usuarios = Usuario.query.all()
    return render_template('logs_auditoria.html', logs=logs, usuarios=usuarios, now=agora)

@app.route('/exportar-logs')
@login_required
@admin_required
def exportar_logs():
    logs = LogAuditoria.query.order_by(LogAuditoria.data_hora.desc()).all()
    
    data = []
    for log in logs:
        dados_formatados = ""
        if log.dados:
            try:
                dados_dict = json.loads(log.dados)
                dados_formatados = json.dumps(dados_dict, ensure_ascii=False, indent=2)
            except:
                dados_formatados = log.dados
        
        data.append({
            'ID': log.id,
            'Data/Hora': log.data_hora.strftime('%d/%m/%Y %H:%M:%S'),
            'Usuário': log.usuario.username if log.usuario else 'Sistema',
            'Nome Usuário': log.usuario.nome if log.usuario and log.usuario.nome else 'N/A',
            'Módulo': log.modulo.upper(),
            'Ação': log.acao,
            'IP': log.ip,
            'Dados': dados_formatados
        })
    
    output = export_to_excel(data, 'logs_auditoria_ro_experience.xlsx', 'Logs de Auditoria')
    
    registrar_log('exportacao_logs', 'logs', {
        'quantidade_logs': len(logs),
        'tipo_exportacao': 'excel'
    })
    
    return send_file(
        output,
        download_name='logs_auditoria_ro_experience.xlsx',
        as_attachment=True,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

@app.route('/pesquisa', methods=['GET', 'POST'])
def pesquisa_publica():
    if request.method == 'POST':
        try:
            print("\n" + "="*60)
            print("🔍 DEBUG - INICIANDO PROCESSAMENTO DA PESQUISA")
            print("="*60)
            
            # DEBUG: Verificar todos os dados recebidos
            print("📦 DADOS RECEBIDOS DO FORMULÁRIO:")
            for key, value in request.form.items():
                print(f"   {key}: {value}")
            
            # NOVAS PERGUNTAS (1-14)
            comunicacao = int(request.form.get('comunicacao'))
            formato_evento = int(request.form.get('formato_evento'))
            alimentacao = int(request.form.get('alimentacao'))
            palestra_reforma = int(request.form.get('palestra_reforma'))
            palestra_estrategia = int(request.form.get('palestra_estrategia'))
            organizacao = int(request.form.get('organizacao'))
            interacao_brother = int(request.form.get('interacao_brother'))
            interacao_canon = int(request.form.get('interacao_canon'))
            interacao_epson = int(request.form.get('interacao_epson'))
            interacao_hp = int(request.form.get('interacao_hp'))
            interacao_konica = int(request.form.get('interacao_konica'))
            interacao_kyocera = int(request.form.get('interacao_kyocera'))
            prazo_entrega = int(request.form.get('prazo_entrega'))
            frete = int(request.form.get('frete'))
            
            comentarios = request.form.get('comentarios', '').strip()
            cnpj = request.form.get('cnpj', '').strip()
            
            print(f"📝 CNPJ RECEBIDO: '{cnpj}'")
            print(f"📏 Tamanho do CNPJ: {len(cnpj)}")
            
            # Validação de todas as questões obrigatórias
            campos_obrigatorios = [
                comunicacao, formato_evento, alimentacao, palestra_reforma, palestra_estrategia,
                organizacao, interacao_brother, interacao_canon, interacao_epson, interacao_hp,
                interacao_konica, interacao_kyocera, prazo_entrega, frete
            ]
            
            if not all(campos_obrigatorios):
                flash('Por favor, responda todas as questões obrigatórias', 'error')
                return render_template('pesquisa_publica.html', enviado=False)
            
            # LÓGICA DE IDENTIFICAÇÃO CORRIGIDA
            anonima = True
            cnpj_identificado = None
            razao_social = None
            
            if cnpj and cnpj.strip():
                cnpj_limpo = ''.join(filter(str.isdigit, cnpj))
                print(f"🔧 CNPJ LIMPO: '{cnpj_limpo}'")
                print(f"📏 Tamanho CNPJ limpo: {len(cnpj_limpo)}")
                
                if len(cnpj_limpo) == 14:
                    # BUSCA 1: Busca exata pelo CNPJ limpo
                    cliente = Cliente.query.filter_by(cnpj=cnpj_limpo).first()
                    print(f"🔎 Busca 1 (CNPJ limpo): {'ENCONTRADO' if cliente else 'NÃO ENCONTRADO'}")
                    
                    if not cliente:
                        # BUSCA 2: Busca com CNPJ formatado
                        cnpj_formatado = normalizar_cnpj(cnpj_limpo)
                        cliente = Cliente.query.filter_by(cnpj=cnpj_formatado).first()
                        print(f"🔎 Busca 2 (CNPJ formatado '{cnpj_formatado}'): {'ENCONTRADO' if cliente else 'NÃO ENCONTRADO'}")
                    
                    if not cliente:
                        # BUSCA 3: Busca em todos os clientes
                        todos_clientes = Cliente.query.all()
                        print(f"🔎 Busca 3 (em {len(todos_clientes)} clientes)")
                        for cli in todos_clientes:
                            cli_cnpj_limpo = ''.join(filter(str.isdigit, cli.cnpj))
                            if cli_cnpj_limpo == cnpj_limpo:
                                cliente = cli
                                print(f"   ✅ Cliente encontrado: {cli.razao_social}")
                                print(f"   📋 CNPJ no banco: '{cli.cnpj}'")
                                break
                    
                    if cliente:
                        print(f"🎯 CLIENTE ENCONTRADO: {cliente.razao_social}")
                        print(f"✅ Check-in: {cliente.checkin_realizado}")
                        print(f"📋 CNPJ no banco: '{cliente.cnpj}'")
                        
                        if cliente.checkin_realizado:
                            anonima = False
                            cnpj_identificado = cliente.cnpj  # USA O CNPJ LIMPO
                            razao_social = cliente.razao_social
                            print(f"🎉 PESQUISA IDENTIFICADA: {razao_social}")
                        else:
                            print("⚠️ Cliente SEM check-in")
                    else:
                        print("❌ Cliente NÃO encontrado")
                        # Debug adicional: mostra alguns clientes
                        alguns_clientes = Cliente.query.filter_by(checkin_realizado=True).limit(3).all()
                        print("📋 Clientes com check-in no banco:")
                        for cli in alguns_clientes:
                            print(f"   - {cli.cnpj} | {cli.razao_social}")
                else:
                    print(f"❌ CNPJ inválido: {len(cnpj_limpo)} dígitos (precisa 14)")
            else:
                print("📝 CNPJ vazio - resposta ANÔNIMA")
            
            print(f"📊 STATUS FINAL: {'ANÔNIMA' if anonima else 'IDENTIFICADA'}")
            print(f"📦 CNPJ que será salvo: {cnpj_identificado}")
            print(f"🏢 Razão Social: {razao_social}")
            print("="*60)
            
            resposta = PesquisaResposta(
                cnpj_identificado=cnpj_identificado,
                razao_social=razao_social,
                
                # NOVAS PERGUNTAS (1-14)
                comunicacao=comunicacao,
                formato_evento=formato_evento,
                alimentacao=alimentacao,
                palestra_reforma=palestra_reforma,
                palestra_estrategia=palestra_estrategia,
                organizacao=organizacao,
                interacao_brother=interacao_brother,
                interacao_canon=interacao_canon,
                interacao_epson=interacao_epson,
                interacao_hp=interacao_hp,
                interacao_konica=interacao_konica,
                interacao_kyocera=interacao_kyocera,
                prazo_entrega=prazo_entrega,
                frete=frete,
                
                comentarios=comentarios if comentarios else None,
                ip=request.remote_addr,
                anonima=anonima
            )
            
            db.session.add(resposta)
            db.session.commit()
            
            registrar_log('pesquisa_respondida', 'pesquisa', {
                'resposta_id': resposta.id,
                'anonima': anonima,
                'cnpj': cnpj_identificado,
                'razao_social': razao_social
            })
            
            return render_template('pesquisa_publica.html', enviado=True)
            
        except Exception as e:
            db.session.rollback()
            print(f"❌ Erro ao salvar pesquisa: {str(e)}")
            import traceback
            traceback.print_exc()
            flash(f'Erro ao enviar pesquisa: {str(e)}', 'error')
            return render_template('pesquisa_publica.html', enviado=False)
    
    return render_template('pesquisa_publica.html', enviado=False)

@app.route('/api/validar-cnpj-pesquisa/', defaults={'cnpj': ''})
@app.route('/api/validar-cnpj-pesquisa/<path:cnpj>')
def api_validar_cnpj_pesquisa(cnpj):
    try:
        from urllib.parse import unquote
        cnpj = unquote(cnpj)
        
        cnpj_limpo = ''.join(filter(str.isdigit, cnpj))
        
        print(f"🔍 Validando CNPJ: {cnpj} -> {cnpj_limpo}")
        
        if len(cnpj_limpo) != 14:
            return jsonify({'valido': False, 'mensagem': 'CNPJ deve ter 14 dígitos'})
        
        # Busca exata pelo CNPJ limpo
        cliente = Cliente.query.filter_by(cnpj=cnpj_limpo).first()
        
        if not cliente:
            # Tenta com CNPJ formatado
            cnpj_formatado = normalizar_cnpj(cnpj_limpo)
            cliente = Cliente.query.filter_by(cnpj=cnpj_formatado).first()
            print(f"🔍 Tentando busca com formato: {cnpj_formatado}")
        
        if not cliente:
            # Busca em todos os clientes (fallback)
            todos_clientes = Cliente.query.all()
            for cli in todos_clientes:
                cli_cnpj_limpo = ''.join(filter(str.isdigit, cli.cnpj))
                if cli_cnpj_limpo == cnpj_limpo:
                    cliente = cli
                    break
        
        if cliente:
            print(f"📋 Cliente encontrado: {cliente.razao_social}")
            print(f"✅ Check-in realizado: {cliente.checkin_realizado}")
            print(f"📝 CNPJ no banco: {cliente.cnpj}")
            
            if cliente.checkin_realizado:
                return jsonify({
                    'valido': True, 
                    'mensagem': f'CNPJ validado - {cliente.razao_social}',
                    'cnpj_correto': cliente.cnpj  # Retorna o CNPJ exato do banco
                })
            else:
                return jsonify({
                    'valido': False, 
                    'mensagem': 'CNPJ encontrado mas não fez check-in no evento'
                })
        else:
            return jsonify({
                'valido': False, 
                'mensagem': 'Esse CNPJ não fez Check-in no evento, use o CNPJ do seu crachá'
            })
            
    except Exception as e:
        print(f"❌ Erro na validação: {str(e)}")
        return jsonify({'valido': False, 'mensagem': f'Erro na validação: {str(e)}'})
    
@app.route('/pesquisa-marketing')
@login_required
@permissao_required('pesquisa_marketing')  # ← Deve ser esta permissão
def pesquisa_marketing():
    return render_template('pesquisa_marketing.html')

@app.route('/api/verificar-cnpj-marketing')
@login_required
def api_verificar_cnpj_marketing():
    """Verifica se CNPJ está qualificado para pesquisa marketing"""
    cnpj = request.args.get('cnpj')
    # resto do código igual...
    try:
        from urllib.parse import unquote
        cnpj = unquote(cnpj)
        
        cnpj_limpo = ''.join(filter(str.isdigit, cnpj))
        
        if len(cnpj_limpo) != 14:
            return jsonify({'valido': False, 'mensagem': 'CNPJ deve ter 14 dígitos'})
        
        # Busca cliente com check-in realizado
        cliente = Cliente.query.filter_by(cnpj=cnpj_limpo, checkin_realizado=True).first()
        
        if not cliente:
            # Tenta com CNPJ formatado
            cnpj_formatado = normalizar_cnpj(cnpj_limpo)
            cliente = Cliente.query.filter_by(cnpj=cnpj_formatado, checkin_realizado=True).first()
        
        if not cliente:
            # Busca em todos os clientes (fallback)
            todos_clientes = Cliente.query.filter_by(checkin_realizado=True).all()
            for cli in todos_clientes:
                cli_cnpj_limpo = ''.join(filter(str.isdigit, cli.cnpj))
                if cli_cnpj_limpo == cnpj_limpo:
                    cliente = cli
                    break
        
        if cliente:
            return jsonify({
                'valido': True, 
                'mensagem': f'CNPJ validado - {cliente.razao_social}',
                'razao_social': cliente.razao_social,
                'responsavel': cliente.responsavel,
                'cnpj_correto': cliente.cnpj
            })
        else:
            return jsonify({
                'valido': False, 
                'mensagem': 'CNPJ não encontrado ou não fez check-in no evento'
            })
            
    except Exception as e:
        return jsonify({'valido': False, 'mensagem': f'Erro na validação: {str(e)}'})


@app.route('/submit_pesquisa_marketing', methods=['POST'])
@login_required
@permissao_required('pesquisa_marketing')
def submit_pesquisa_marketing():
    try:
        data = request.get_json()
        
        # CNPJ é OBRIGATÓRIO para pesquisa marketing
        cnpj = data.get('cnpj', '').strip()
        if not cnpj:
            return jsonify({'success': False, 'message': 'CNPJ é obrigatório para esta pesquisa'})
        
        cnpj_limpo = ''.join(filter(str.isdigit, cnpj))
        
        # Verificar novamente se o cliente existe e fez check-in
        cliente = Cliente.query.filter_by(cnpj=cnpj_limpo, checkin_realizado=True).first()
        if not cliente:
            # Tenta com CNPJ formatado
            cnpj_formatado = normalizar_cnpj(cnpj_limpo)
            cliente = Cliente.query.filter_by(cnpj=cnpj_formatado, checkin_realizado=True).first()
        
        if not cliente:
            return jsonify({'success': False, 'message': 'CNPJ não validado ou não fez check-in no evento'})
        
        # Processar benefícios de engajamento (questão 2 - múltipla escolha)
        beneficio_engajamento = data.get('beneficio_engajamento', [])
        if isinstance(beneficio_engajamento, list) and len(beneficio_engajamento) > 2:
            return jsonify({'success': False, 'message': 'Selecione no máximo 2 opções na questão 2'})

        # Validar campos obrigatórios
        if not data.get('posicionamento'):
            return jsonify({'success': False, 'message': 'A questão 1 (posicionamento) é obrigatória'})
        
        if not data.get('beneficio_preferido'):
            return jsonify({'success': False, 'message': 'A questão 3 (benefício preferido) é obrigatória'})
        
        if not data.get('valor_parceiro'):
            return jsonify({'success': False, 'message': 'A questão 10 (valor do parceiro) é obrigatória'})

        # Criar a pesquisa com valores padrão para campos opcionais
        pesquisa = PesquisaMarketing(
            cnpj_identificado=cliente.cnpj,
            razao_social=cliente.razao_social,
            usuario_responsavel=cliente.responsavel,
            
            # Questão 1 (OBRIGATÓRIA)
            posicionamento=data.get('posicionamento'),
            
            # Questão 2 (OPCIONAL)
            beneficio_engajamento=json.dumps(beneficio_engajamento) if beneficio_engajamento else None,
            
            # Questão 3 (OBRIGATÓRIA)
            beneficio_preferido=data.get('beneficio_preferido'),
            
            # Questão 4 (OPCIONAL) - usar get com valor padrão 0
            margem_lucro=int(data.get('margem_lucro', 0)),
            qualidade_produtos=int(data.get('qualidade_produtos', 0)),
            suporte_comercial=int(data.get('suporte_comercial', 0)),
            condicoes_comerciais=int(data.get('condicoes_comerciais', 0)),
            reconhecimento_marca=int(data.get('reconhecimento_marca', 0)),
            velocidade_resposta=int(data.get('velocidade_resposta', 0)),
            facilidade_pedidos=int(data.get('facilidade_pedidos', 0)),
            
            # Questões 5-9 (OPCIONAIS)
            dificuldade_participacao=data.get('dificuldade_participacao'),
            tipo_campanha_impacto=data.get('tipo_campanha_impacto'),
            beneficio_venda=data.get('beneficio_venda'),
            aumento_volume=data.get('aumento_volume'),
            competitividade=data.get('competitividade'),
            
            # Questão 10 (OBRIGATÓRIA)
            valor_parceiro=data.get('valor_parceiro'),
            
            # Campos "Outro" (OPCIONAIS)
            outro_questao2=data.get('outro_questao2'),
            outro_questao3=data.get('outro_questao3'),
            outro_questao5=data.get('outro_questao5'),
            outro_questao6=data.get('outro_questao6'),
            outro_questao8=data.get('outro_questao8'),
            outro_questao10=data.get('outro_questao10'),
            
            comentarios_gerais=data.get('comentarios_gerais'),
            ip=request.remote_addr
        )
        
        db.session.add(pesquisa)
        db.session.commit()
        
        registrar_log('pesquisa_marketing_respondida', 'pesquisa_mkt', {
            'pesquisa_id': pesquisa.id,
            'cnpj': cliente.cnpj,
            'razao_social': cliente.razao_social,
            'responsavel': cliente.responsavel
        })
        
        return jsonify({'success': True, 'message': 'Pesquisa de marketing enviada com sucesso!'})
        
    except Exception as e:
        db.session.rollback()
        print(f"❌ Erro ao salvar pesquisa marketing: {str(e)}")
        return jsonify({'success': False, 'message': f'Erro ao enviar pesquisa: {str(e)}'})
@app.route('/fix-pesquisa-table')
def fix_pesquisa_table():
    """Corrige a tabela pesquisa_marketing para permitir valores nulos"""
    try:
        with db.engine.connect() as conn:
            # Altera as colunas para permitir NULL
            conn.execute(db.text("""
                ALTER TABLE pesquisa_marketing 
                ALTER COLUMN dificuldade_participacao DROP NOT NULL,
                ALTER COLUMN tipo_campanha_impacto DROP NOT NULL,
                ALTER COLUMN beneficio_venda DROP NOT NULL,
                ALTER COLUMN aumento_volume DROP NOT NULL,
                ALTER COLUMN competitividade DROP NOT NULL;
            """))
            conn.commit()
        
        return jsonify({'success': True, 'message': 'Tabela corrigida com sucesso!'})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})
    
@app.route('/relatorio-pesquisas-marketing')
@login_required
@permissao_required('relatorio_pesquisas_mkt')
def relatorio_pesquisas_marketing():
    pesquisas = PesquisaMarketing.query.order_by(PesquisaMarketing.data_resposta.desc()).all()
    
    total_pesquisas = len(pesquisas)
    pesquisas_identificadas = len([p for p in pesquisas if p.cnpj_identificado])
    
    # Calcular médias para questão 4
    campos_questao4 = ['margem_lucro', 'qualidade_produtos', 'suporte_comercial', 'condicoes_comerciais', 
                      'reconhecimento_marca', 'velocidade_resposta', 'facilidade_pedidos']
    
    medias = {}
    for campo in campos_questao4:
        if total_pesquisas > 0:
            valores = [getattr(p, campo) for p in pesquisas if getattr(p, campo) > 0]
            medias[campo] = sum(valores) / len(valores) if valores else 0
        else:
            medias[campo] = 0
    
    # Estatísticas para questões de múltipla escolha
    stats = {
        'posicionamento': {},
        'beneficio_preferido': {},
        'dificuldade_participacao': {},
        'tipo_campanha_impacto': {},
        'aumento_volume': {},
        'competitividade': {},
        'valor_parceiro': {},
        'beneficio_engajamento': {}
    }
    
    for pesquisa in pesquisas:
        # Questão 1
        stats['posicionamento'][pesquisa.posicionamento] = stats['posicionamento'].get(pesquisa.posicionamento, 0) + 1
        
        # Questão 3
        stats['beneficio_preferido'][pesquisa.beneficio_preferido] = stats['beneficio_preferido'].get(pesquisa.beneficio_preferido, 0) + 1
        
        # Questão 5
        stats['dificuldade_participacao'][pesquisa.dificuldade_participacao] = stats['dificuldade_participacao'].get(pesquisa.dificuldade_participacao, 0) + 1
        
        # Questão 6
        stats['tipo_campanha_impacto'][pesquisa.tipo_campanha_impacto] = stats['tipo_campanha_impacto'].get(pesquisa.tipo_campanha_impacto, 0) + 1
        
        # Questão 8
        stats['aumento_volume'][pesquisa.aumento_volume] = stats['aumento_volume'].get(pesquisa.aumento_volume, 0) + 1
        
        # Questão 9
        stats['competitividade'][pesquisa.competitividade] = stats['competitividade'].get(pesquisa.competitividade, 0) + 1
        
        # Questão 10
        stats['valor_parceiro'][pesquisa.valor_parceiro] = stats['valor_parceiro'].get(pesquisa.valor_parceiro, 0) + 1
        
        # Questão 2 (múltipla)
        beneficios = json.loads(pesquisa.beneficio_engajamento) if pesquisa.beneficio_engajamento else []
        for beneficio in beneficios:
            stats['beneficio_engajamento'][beneficio] = stats['beneficio_engajamento'].get(beneficio, 0) + 1
    
    return render_template('relatorio_pesquisas_marketing.html',
                         pesquisas=pesquisas,
                         total_pesquisas=total_pesquisas,
                         pesquisas_identificadas=pesquisas_identificadas,
                         medias=medias,
                         stats=stats)

@app.route('/exportar-pesquisas-marketing')
@login_required
@permissao_required('exportacao')
def exportar_pesquisas_marketing():
    pesquisas = PesquisaMarketing.query.order_by(PesquisaMarketing.data_resposta.desc()).all()
    
    data = []
    for pesquisa in pesquisas:
        beneficios_engajamento = json.loads(pesquisa.beneficio_engajamento) if pesquisa.beneficio_engajamento else []
        
        data.append({
            'ID': pesquisa.id,
            'Data/Hora': pesquisa.data_resposta.strftime('%d/%m/%Y %H:%M'),
            'CNPJ': pesquisa.cnpj_identificado or 'N/A',
            'Razão Social': pesquisa.razao_social or 'N/A',
            'Responsável': pesquisa.usuario_responsavel or 'N/A',
            'Posicionamento': pesquisa.posicionamento,
            'Benefícios Engajamento': ', '.join(beneficios_engajamento),
            'Outro Benefício Engajamento': pesquisa.outro_questao2 or 'N/A',
            'Benefício Preferido': pesquisa.beneficio_preferido,
            'Outro Benefício Preferido': pesquisa.outro_questao3 or 'N/A',
            'Margem Lucro (1-5)': pesquisa.margem_lucro,
            'Qualidade Produtos (1-5)': pesquisa.qualidade_produtos,
            'Suporte Comercial (1-5)': pesquisa.suporte_comercial,
            'Condições Comerciais (1-5)': pesquisa.condicoes_comerciais,
            'Reconhecimento Marca (1-5)': pesquisa.reconhecimento_marca,
            'Velocidade Resposta (1-5)': pesquisa.velocidade_resposta,
            'Facilidade Pedidos (1-5)': pesquisa.facilidade_pedidos,
            'Dificuldade Participação': pesquisa.dificuldade_participacao,
            'Outra Dificuldade': pesquisa.outro_questao5 or 'N/A',
            'Tipo Campanha Impacto': pesquisa.tipo_campanha_impacto,
            'Outro Tipo Campanha': pesquisa.outro_questao6 or 'N/A',
            'Benefício Venda': pesquisa.beneficio_venda or 'N/A',
            'Aumento Volume': pesquisa.aumento_volume,
            'Outro Aumento Volume': pesquisa.outro_questao8 or 'N/A',
            'Competitividade': pesquisa.competitividade,
            'Valor Parceiro': pesquisa.valor_parceiro,
            'Outro Valor Parceiro': pesquisa.outro_questao10 or 'N/A',
            'Comentários Gerais': pesquisa.comentarios_gerais or 'N/A',
            'IP': pesquisa.ip
        })
    
    output = export_to_excel(data, 'pesquisas_marketing.xlsx', 'Pesquisas Marketing')
    
    registrar_log('exportacao_pesquisas_marketing', 'exportacao', {
        'quantidade_pesquisas': len(pesquisas)
    })
    
    return send_file(output,
                    download_name='pesquisas_marketing_ro_experience.xlsx',
                    as_attachment=True,
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    
    
    
@app.route('/relatorio-pesquisas')
@login_required
@permissao_required('relatorio_pesquisas')  # ← Deve ser esta permissão
def relatorio_pesquisas():
    pesquisas = PesquisaResposta.query.order_by(PesquisaResposta.data_resposta.desc()).all()
    
    total_pesquisas = len(pesquisas)
    pesquisas_identificadas = len([p for p in pesquisas if not p.anonima])
    pesquisas_anonimas = len([p for p in pesquisas if p.anonima])
    
    # Calcular médias para todas as NOVAS questões
    campos = [
        'comunicacao', 'formato_evento', 'alimentacao', 'palestra_reforma', 'palestra_estrategia',
        'organizacao', 'interacao_brother', 'interacao_canon', 'interacao_epson', 'interacao_hp',
        'interacao_konica', 'interacao_kyocera', 'prazo_entrega', 'frete'
    ]
    
    medias = {}
    for campo in campos:
        if total_pesquisas > 0:
            medias[campo] = sum(getattr(p, campo) for p in pesquisas) / total_pesquisas
        else:
            medias[campo] = 0
    
    return render_template('relatorio_pesquisas.html',
                         pesquisas=pesquisas,
                         total_pesquisas=total_pesquisas,
                         pesquisas_identificadas=pesquisas_identificadas,
                         pesquisas_anonimas=pesquisas_anonimas,
                         medias=medias)

@app.route('/exportar-pesquisas')
@login_required
@permissao_required('exportacao')
def exportar_pesquisas():
    pesquisas = PesquisaResposta.query.order_by(PesquisaResposta.data_resposta.desc()).all()
    
    data = []
    for pesquisa in pesquisas:
        data.append({
            'ID': pesquisa.id,
            'Data/Hora': pesquisa.data_resposta.strftime('%d/%m/%Y %H:%M'),
            'Tipo': 'Anônima' if pesquisa.anonima else 'Identificada',
            'CNPJ': pesquisa.cnpj_identificado or 'N/A',
            'Razão Social': pesquisa.razao_social or 'N/A',
            
            # NOVOS CAMPOS (1-5)
            'Comunicação (1-5)': pesquisa.comunicacao,
            'Formato Evento (1-5)': pesquisa.formato_evento,
            'Alimentação (1-5)': pesquisa.alimentacao,
            'Palestra Reforma (1-5)': pesquisa.palestra_reforma,
            'Palestra Estratégia (1-5)': pesquisa.palestra_estrategia,
            
            # CAMPOS EXISTENTES ATUALIZADOS (6-14)
            'Organização (1-5)': pesquisa.organizacao,
            'Interação Brother (1-5)': pesquisa.interacao_brother,
            'Interação Canon (1-5)': pesquisa.interacao_canon,
            'Interação Epson (1-5)': pesquisa.interacao_epson,
            'Interação HP (1-5)': pesquisa.interacao_hp,
            'Interação Konica (1-5)': pesquisa.interacao_konica,
            'Interação Kyocera (1-5)': pesquisa.interacao_kyocera,
            'Prazo Entrega (1-5)': pesquisa.prazo_entrega,
            'Frete (1-5)': pesquisa.frete,
            
            'Comentários': pesquisa.comentarios or 'Nenhum',
            'IP': pesquisa.ip
        })
    
    output = export_to_excel(data, 'pesquisas_satisfacao.xlsx', 'Pesquisas')
    
    registrar_log('exportacao_pesquisas', 'exportacao', {
        'quantidade_pesquisas': len(pesquisas)
    })
    
    return send_file(
        output,
        download_name='pesquisas_satisfacao_ro_experience.xlsx',
        as_attachment=True,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

# Arquivo: test_postgres.py
import psycopg2
import os

def testar_permissoes():
    try:
        conn = psycopg2.connect(
            host="square-cloud-db-4d0ca60ac1a54ad48adf5608996c6a48.squareweb.app",
            port=7091,
            user="squarecloud",
            password="5W3Ww67llyHrBmcutvyL5xXO",
            database="postgres",
            sslmode="require",
            sslrootcert="ca-certificate.crt",
            sslcert="certificate.pem", 
            sslkey="private-key.key"
        )
        
        cursor = conn.cursor()
        
        # Testa permissões básicas
        cursor.execute("SELECT current_user, current_database()")
        print(f"✅ Conectado como: {cursor.fetchone()}")
        
        # Testa criar schema
        try:
            cursor.execute("CREATE SCHEMA IF NOT EXISTS test_permissions")
            print("✅ Pode criar schemas")
        except Exception as e:
            print(f"❌ NÃO pode criar schemas: {e}")
            
        # Testa criar tabela
        try:
            cursor.execute("CREATE TABLE test_table (id SERIAL PRIMARY KEY, name TEXT)")
            print("✅ Pode criar tabelas")
        except Exception as e:
            print(f"❌ NÃO pode criar tabelas: {e}")
            
        conn.close()
        
    except Exception as e:
        print(f"❌ Erro de conexão: {e}")

 
def criar_banco_se_nao_existir(app):
    """
    Cria o banco de dados alvo ('dbexperience') se não existir, conectando-se 
    ao banco padrão 'postgres' com as mesmas configurações SSL e credenciais.
    """
    full_uri = app.config['SQLALCHEMY_DATABASE_URI']
    
    # 1. Isola o nome do banco de dados alvo ('dbexperience')
    db_name_and_query = full_uri.split('/')[-1]
    DB_NAME = db_name_and_query.split('?')[0]

    # 2. Constrói a URI para a conexão temporária no banco 'postgres'
    # Substitui o nome do banco alvo pelo banco default 'postgres'
    temp_uri = full_uri.replace(f'/{DB_NAME}', '/postgres') 
    
    # 3. Remove o prefixo '+psycopg2' que só é usado pelo SQLAlchemy
    temp_uri = temp_uri.replace('+psycopg2', '')
    
    conn = None
    try:
        print(f"🔧 Tentando conectar ao banco default ('postgres') para criar '{DB_NAME}'...")
        
        # Conexão usando a URI completa, que inclui todos os parâmetros SSL
        conn = psycopg2.connect(temp_uri) 
        
        conn.autocommit = True
        cursor = conn.cursor()
        
        # Verifica se o banco de dados alvo existe
        cursor.execute(f"SELECT 1 FROM pg_database WHERE datname = %s", (DB_NAME,))
        exists = cursor.fetchone()
        
        if not exists:
            # Cria o banco de dados
            print(f"🔧 Criando banco de dados '{DB_NAME}'...")
            cursor.execute(f"CREATE DATABASE {DB_NAME}")
            print(f"✅ Banco de dados '{DB_NAME}' criado com sucesso!")
        else:
            print(f"✅ Banco de dados '{DB_NAME}' já existe no servidor.")
            
        cursor.close()

    except psycopg2.OperationalError as e:
        # Se falhar aqui, o erro é de conexão ou certificado
        print(f"❌ Erro crítico ao criar/verificar o banco de dados: {e}")
        raise Exception("Conexão ao DB default falhou. Verifique as credenciais, certificados e permissões.") from e
        
    except Exception as e:
        print(f"❌ Erro inesperado ao configurar o banco de dados: {e}")
        raise

    finally:
        if conn:
            conn.close()  

@app.route('/importar-vendas-evento', methods=['GET', 'POST'])
@login_required
@permissao_required('importacao_vendas')
def importar_vendas_evento():
    if request.method == 'POST':
        if 'file' not in request.files:
            flash('Nenhum arquivo selecionado', 'error')
            return redirect(request.url)
        
        file = request.files['file']
        if file.filename == '':
            flash('Nenhum arquivo selecionado', 'error')
            return redirect(request.url)
        
        if file and file.filename.endswith(('.xlsx', '.xls')):
            try:
                data = read_excel_file(file)
                
                # VERIFICAÇÃO DE COLUNAS ATUALIZADA
                colunas_necessarias = ['NF', 'DATA_EMISSAO', 'CLIENTE_NOME', 'VENDEDOR', 'EQUIPE', 
                                      'DESCRICAO_PRODUTO', 'MARCA', 'VALOR_PRODUTOS', 'QTD']  # QTD no lugar de QUANTIDADE
                
                if data and data[0]:
                    colunas_arquivo = [key.upper() for key in data[0].keys()]
                    colunas_faltando = [c for c in colunas_necessarias if c not in colunas_arquivo]
                    
                    if colunas_faltando:
                        flash(f'Colunas faltando no arquivo: {", ".join(colunas_faltando)}', 'error')
                        return redirect(request.url)
                else:
                    flash('Arquivo vazio ou formato inválido', 'error')
                    return redirect(request.url)
                
                vendas_importadas = 0
                erros = []
                
                for i, row in enumerate(data, 1):
                    try:
                        # NOVO: Obter número da NF
                        numero_nf = str(row.get('NF', '')).strip()
                        if not numero_nf:
                            numero_nf = f"SEM_NF_{i}"
                        
                        # Converter data
                        data_str = str(row.get('DATA_EMISSAO', ''))
                        if data_str:
                            try:
                                if '-' in data_str:
                                    data_emissao = datetime.strptime(data_str, '%Y-%m-%d').date()
                                elif '/' in data_str:
                                    data_emissao = datetime.strptime(data_str, '%d/%m/%Y').date()
                                else:
                                    data_emissao = datetime.fromtimestamp(float(data_str)).date()
                            except:
                                data_emissao = agora().date()
                        else:
                            data_emissao = agora().date()
                        
                        # Obter valores
                        cliente_nome = str(row.get('CLIENTE_NOME', '')).strip()
                        vendedor = str(row.get('VENDEDOR', '')).strip()
                        equipe = str(row.get('EQUIPE', '')).strip()
                        descricao = str(row.get('DESCRICAO_PRODUTO', '')).strip()
                        marca = str(row.get('MARCA', '')).strip()
                        familia = str(row.get('FAMILIA', '')).strip() if row.get('FAMILIA') else None
                        
                        # Converter valores numéricos
                        try:
                            valor_produtos = float(str(row.get('VALOR_PRODUTOS', '0')).replace(',', '.'))
                        except:
                            valor_produtos = 0.0
                        
                        try:
                            quantidade = int(float(str(row.get('QTD', '1'))))  # QTD no lugar de QUANTIDADE
                        except:
                            quantidade = 1
                        
                        # Valor total = valor_produtos (já é o total)
                        valor_total = valor_produtos
                        
                        # SEMPRE CRIAR NOVO REGISTRO - SEM VERIFICAR DUPLICATAS
                        venda = VendaEvento(
                            numero_nf=numero_nf,  # NOVO CAMPO
                            data_emissao=data_emissao,
                            cliente_nome=cliente_nome,
                            vendedor=vendedor,
                            equipe=equipe,
                            descricao_produto=descricao,
                            marca=marca,
                            valor_produtos=valor_produtos,
                            quantidade=quantidade,
                            familia=familia,
                            valor_total=valor_total,
                            importado_por=session.get('nome', 'Sistema')
                        )
                        db.session.add(venda)
                        vendas_importadas += 1
                        
                        # Log a cada 100 registros para monitorar progresso
                        if vendas_importadas % 100 == 0:
                            print(f"📊 Processados {vendas_importadas} registros...")
                            
                    except Exception as e:
                        erros.append(f'Linha {i}: {str(e)}')
                        continue
                
                db.session.commit()
                
                # Registrar log
                registrar_log('importacao_vendas_evento', 'importacao_vendas', {
                    'vendas_importadas': vendas_importadas,
                    'total_linhas': len(data),
                    'erros': len(erros),
                    'arquivo': file.filename
                })
                
                mensagem = f'✅ Importação concluída!<br>'
                mensagem += f'📊 {vendas_importadas} vendas importadas<br>'
                mensagem += f'📈 Total processado: {len(data)} linhas'
                
                if erros:
                    mensagem += f'<br>⚠️ {len(erros)} erros encontrados'
                    flash(mensagem, 'warning')
                else:
                    flash(mensagem, 'success')
                
                return redirect(url_for('analise_vendas'))
                
            except Exception as e:
                flash(f'Erro na importação: {str(e)}', 'error')
                return redirect(request.url)
    
    # GET - mostrar página de importação
    total_vendas = VendaEvento.query.count()
    total_clientes = db.session.query(VendaEvento.cliente_nome).distinct().count()
    total_vendedores = db.session.query(VendaEvento.vendedor).distinct().count()
    
    return render_template('importar_vendas.html',
                         total_vendas=total_vendas,
                         total_clientes=total_clientes,
                         total_vendedores=total_vendedores)

@app.route('/analise-vendas')
@login_required
@permissao_required('analise_vendas')
def analise_vendas():
    """Dashboard de análise de vendas do evento"""
    
    # Obter dados para filtros
    clientes = db.session.query(VendaEvento.cliente_nome).distinct().order_by(VendaEvento.cliente_nome).all()
    vendedores = db.session.query(VendaEvento.vendedor).distinct().order_by(VendaEvento.vendedor).all()
    equipes = db.session.query(VendaEvento.equipe).distinct().order_by(VendaEvento.equipe).all()
    marcas = db.session.query(VendaEvento.marca).distinct().order_by(VendaEvento.marca).all()
    familias = db.session.query(VendaEvento.familia).distinct().filter(VendaEvento.familia.isnot(None)).order_by(VendaEvento.familia).all()
    
    # Datas mínima e máxima
    min_date = db.session.query(db.func.min(VendaEvento.data_emissao)).scalar()
    max_date = db.session.query(db.func.max(VendaEvento.data_emissao)).scalar()
    
    return render_template('analise_vendas.html',
                         clientes=[c[0] for c in clientes],
                         vendedores=[v[0] for v in vendedores],
                         equipes=[e[0] for e in equipes],
                         marcas=[m[0] for m in marcas],
                         familias=[f[0] for f in familias],
                         min_date=min_date,
                         max_date=max_date)

@app.route('/api/vendas-filtradas', methods=['POST'])
@login_required
def api_vendas_filtradas():
    """API para obter vendas com filtros - ATUALIZADA"""
    data = request.get_json()
    
    # Construir query base
    query = VendaEvento.query
    
    # Aplicar filtros
    if data.get('cliente') and data['cliente'] != 'todos':
        query = query.filter(VendaEvento.cliente_nome == data['cliente'])
    
    if data.get('vendedor') and data['vendedor'] != 'todos':
        query = query.filter(VendaEvento.vendedor == data['vendedor'])
    
    if data.get('equipe') and data['equipe'] != 'todos':
        query = query.filter(VendaEvento.equipe == data['equipe'])
    
    if data.get('marca') and data['marca'] != 'todos':
        query = query.filter(VendaEvento.marca == data['marca'])
    
    if data.get('familia') and data['familia'] != 'todos':
        query = query.filter(VendaEvento.familia == data['familia'])
    
    if data.get('data_inicio'):
        try:
            data_inicio = datetime.strptime(data['data_inicio'], '%Y-%m-%d').date()
            query = query.filter(VendaEvento.data_emissao >= data_inicio)
        except:
            pass
    
    if data.get('data_fim'):
        try:
            data_fim = datetime.strptime(data['data_fim'], '%Y-%m-%d').date()
            query = query.filter(VendaEvento.data_emissao <= data_fim)
        except:
            pass
    
    from sqlalchemy import func
    
    # Calcular totais ANTES da paginação
    total_vendas = query.count()
    total_valor = query.with_entities(func.sum(VendaEvento.valor_total)).scalar() or 0
    total_quantidade = query.with_entities(func.sum(VendaEvento.quantidade)).scalar() or 0
    
    # Ordenar
    sort_by = data.get('sort_by', 'data_emissao')
    sort_order = data.get('sort_order', 'desc')
    
    if sort_order == 'desc':
        query = query.order_by(db.desc(getattr(VendaEvento, sort_by, 'data_emissao')))
    else:
        query = query.order_by(getattr(VendaEvento, sort_by, 'data_emissao'))
    
    # Paginação
    page = data.get('page', 1)
    per_page = data.get('per_page', 50)
    
    pagination = query.paginate(page=page, per_page=per_page, error_out=False)
    vendas = pagination.items
    
    result = []
    for venda in vendas:
        # Calcular valor unitário (valor_total / quantidade)
        valor_unitario = venda.valor_total / venda.quantidade if venda.quantidade > 0 else venda.valor_total
        
        result.append({
            'id': venda.id,
            'numero_nf': venda.numero_nf,
            'data_emissao': venda.data_emissao.strftime('%d/%m/%Y'),
            'cliente_nome': venda.cliente_nome,
            'vendedor': venda.vendedor,
            'equipe': venda.equipe,
            'descricao_produto': venda.descricao_produto,
            'marca': venda.marca,
            'valor_produtos': float(venda.valor_produtos),
            'quantidade': venda.quantidade,
            'valor_unitario': float(valor_unitario),
            'familia': venda.familia or 'N/A',
            'valor_total': float(venda.valor_total)
        })
    
    return jsonify({
        'vendas': result,
        'total_vendas': total_vendas,
        'total_valor': float(total_valor),
        'total_quantidade': int(total_quantidade),
        'page': pagination.page,
        'pages': pagination.pages,
        'per_page': pagination.per_page
    })


@app.route('/api/metricas-vendas')
@login_required
def api_metricas_vendas():
    """API para métricas gerais de vendas - CONSIDERANDO NF"""
    
    # TOTAIS CONSIDERANDO NF (NOTAS FISCAIS)
    total_nfs = db.session.query(db.func.count(db.distinct(VendaEvento.numero_nf))).scalar() or 0
    total_vendas_itens = VendaEvento.query.count()
    total_valor = db.session.query(db.func.sum(VendaEvento.valor_total)).scalar() or 0
    total_quantidade = db.session.query(db.func.sum(VendaEvento.quantidade)).scalar() or 0
    
    # Calcular valores médios
    valor_medio_item = total_valor / total_quantidade if total_quantidade > 0 else 0
    valor_medio_nf = total_valor / total_nfs if total_nfs > 0 else 0
    
    # CORREÇÃO: Calcular média de itens por NF usando subquery corretamente
    from sqlalchemy import select, func
    
    # Subquery para contar itens por NF
    subquery_itens = db.session.query(
        VendaEvento.numero_nf,
        func.count(VendaEvento.id).label('total_itens')
    ).group_by(VendaEvento.numero_nf).subquery()
    
    # Subquery para somar valor por NF
    subquery_valor = db.session.query(
        VendaEvento.numero_nf,
        func.sum(VendaEvento.valor_total).label('valor_total')
    ).group_by(VendaEvento.numero_nf).subquery()
    
    # Média de itens por NF
    media_itens_por_nf = db.session.query(
        func.avg(subquery_itens.c.total_itens)
    ).scalar() or 0
    
    # Média de valor por NF
    media_valor_por_nf = db.session.query(
        func.avg(subquery_valor.c.valor_total)
    ).scalar() or 0
    
    # Maior venda (por item)
    maior_venda_item = VendaEvento.query.order_by(VendaEvento.valor_total.desc()).first()
    
    # Maior NF (por valor total)
    maior_nf = db.session.query(
        VendaEvento.numero_nf,
        func.sum(VendaEvento.valor_total).label('valor_total_nf'),
        func.count(VendaEvento.id).label('itens_nf'),
        func.min(VendaEvento.cliente_nome).label('cliente')
    ).group_by(VendaEvento.numero_nf).order_by(func.sum(VendaEvento.valor_total).desc()).first()
    
    # Cliente que mais comprou (por NF)
    cliente_maior_nfs = db.session.query(
        VendaEvento.cliente_nome,
        func.count(func.distinct(VendaEvento.numero_nf)).label('total_nfs'),
        func.sum(VendaEvento.valor_total).label('total_valor'),
        func.sum(VendaEvento.quantidade).label('total_quantidade')
    ).group_by(VendaEvento.cliente_nome).order_by(func.count(func.distinct(VendaEvento.numero_nf)).desc()).first()
    
    # Cliente que mais comprou (por valor)
    cliente_maior_valor = db.session.query(
        VendaEvento.cliente_nome,
        func.sum(VendaEvento.valor_total).label('total_valor'),
        func.count(func.distinct(VendaEvento.numero_nf)).label('total_nfs')
    ).group_by(VendaEvento.cliente_nome).order_by(func.sum(VendaEvento.valor_total).desc()).first()
    
    # Vendedor com mais NFs
    vendedor_maior_nfs = db.session.query(
        VendaEvento.vendedor,
        func.count(func.distinct(VendaEvento.numero_nf)).label('total_nfs'),
        func.sum(VendaEvento.valor_total).label('total_valor'),
        func.count(func.distinct(VendaEvento.cliente_nome)).label('clientes_atendidos')
    ).group_by(VendaEvento.vendedor).order_by(func.count(func.distinct(VendaEvento.numero_nf)).desc()).first()
    
    # Vendedor com maior valor
    vendedor_maior_valor = db.session.query(
        VendaEvento.vendedor,
        func.sum(VendaEvento.valor_total).label('total_valor'),
        func.count(func.distinct(VendaEvento.numero_nf)).label('total_nfs')
    ).group_by(VendaEvento.vendedor).order_by(func.sum(VendaEvento.valor_total).desc()).first()
    
    # Equipe com mais NFs
    equipe_maior_nfs = db.session.query(
        VendaEvento.equipe,
        func.count(func.distinct(VendaEvento.numero_nf)).label('total_nfs'),
        func.sum(VendaEvento.valor_total).label('total_valor')
    ).group_by(VendaEvento.equipe).order_by(func.count(func.distinct(VendaEvento.numero_nf)).desc()).first()
    
    # Marcas por quantidade de NFs
    marcas_por_nf = db.session.query(
        VendaEvento.marca,
        func.count(func.distinct(VendaEvento.numero_nf)).label('total_nfs'),
        func.sum(VendaEvento.valor_total).label('total_valor'),
        func.sum(VendaEvento.quantidade).label('total_quantidade')
    ).group_by(VendaEvento.marca).order_by(func.count(func.distinct(VendaEvento.numero_nf)).desc()).limit(10).all()
    
    # NFs por dia
    nfs_por_dia = db.session.query(
        VendaEvento.data_emissao,
        func.count(func.distinct(VendaEvento.numero_nf)).label('total_nfs'),
        func.sum(VendaEvento.valor_total).label('total_valor')
    ).group_by(VendaEvento.data_emissao).order_by(VendaEvento.data_emissao).all()
    
    return jsonify({
        'totais_por_nf': {
            'total_nfs': total_nfs,
            'total_itens': total_vendas_itens,
            'total_valor': float(total_valor),
            'total_quantidade': int(total_quantidade),
            'media_itens_por_nf': float(media_itens_por_nf),
            'media_valor_por_nf': float(media_valor_por_nf),
            'media_valor_por_item': float(valor_medio_item)
        },
        'maior_nf': {
            'numero_nf': maior_nf[0] if maior_nf else None,
            'valor_total': float(maior_nf[1]) if maior_nf else 0,
            'itens': maior_nf[2] if maior_nf else 0,
            'cliente': maior_nf[3] if maior_nf else None
        },
        'maior_venda_item': {
            'cliente': maior_venda_item.cliente_nome if maior_venda_item else None,
            'valor': float(maior_venda_item.valor_total) if maior_venda_item else 0,
            'descricao': maior_venda_item.descricao_produto if maior_venda_item else None,
            'nf': maior_venda_item.numero_nf if maior_venda_item else None,
            'quantidade': maior_venda_item.quantidade if maior_venda_item else 0
        },
        'clientes': {
            'maior_nfs': {
                'nome': cliente_maior_nfs[0] if cliente_maior_nfs else None,
                'nfs': cliente_maior_nfs[1] if cliente_maior_nfs else 0,
                'valor': float(cliente_maior_nfs[2]) if cliente_maior_nfs else 0,
                'quantidade': cliente_maior_nfs[3] if cliente_maior_nfs else 0
            },
            'maior_valor': {
                'nome': cliente_maior_valor[0] if cliente_maior_valor else None,
                'valor': float(cliente_maior_valor[1]) if cliente_maior_valor else 0,
                'nfs': cliente_maior_valor[2] if cliente_maior_valor else 0
            }
        },
        'vendedores': {
            'maior_nfs': {
                'nome': vendedor_maior_nfs[0] if vendedor_maior_nfs else None,
                'nfs': vendedor_maior_nfs[1] if vendedor_maior_nfs else 0,
                'valor': float(vendedor_maior_nfs[2]) if vendedor_maior_nfs else 0,
                'clientes': vendedor_maior_nfs[3] if vendedor_maior_nfs else 0
            },
            'maior_valor': {
                'nome': vendedor_maior_valor[0] if vendedor_maior_valor else None,
                'valor': float(vendedor_maior_valor[1]) if vendedor_maior_valor else 0,
                'nfs': vendedor_maior_valor[2] if vendedor_maior_valor else 0
            }
        },
        'equipes': {
            'maior_nfs': {
                'nome': equipe_maior_nfs[0] if equipe_maior_nfs else None,
                'nfs': equipe_maior_nfs[1] if equipe_maior_nfs else 0,
                'valor': float(equipe_maior_nfs[2]) if equipe_maior_nfs else 0
            }
        },
        'marcas_por_nf': [
            {
                'marca': marca,
                'nfs': int(total_nfs),
                'valor': float(valor),
                'quantidade': int(quantidade),
                'percentual_nfs': float(total_nfs / total_nfs * 100) if total_nfs > 0 else 0,
                'percentual_valor': float(valor / total_valor * 100) if total_valor > 0 else 0
            }
            for marca, total_nfs, valor, quantidade in marcas_por_nf
        ],
        'nfs_por_dia': [
            {
                'data': data.strftime('%d/%m/%Y'),
                'nfs': int(total_nfs),
                'valor': float(valor)
            }
            for data, total_nfs, valor in nfs_por_dia
        ]
    })

@app.route('/api/capilaridade-vendas')
@login_required
def api_capilaridade_vendas():
    """API para análise de capilaridade por NF, cliente, vendedor e marca"""
    
    from sqlalchemy import func
    
    # 1. CAPILARIDADE POR CLIENTE (REVENDA)
    capilaridade_cliente = db.session.query(
        VendaEvento.cliente_nome,
        func.count(func.distinct(VendaEvento.numero_nf)).label('total_nfs'),
        func.count(VendaEvento.id).label('total_itens'),
        func.sum(VendaEvento.valor_total).label('valor_total'),
        func.sum(VendaEvento.quantidade).label('quantidade_total'),
        func.count(func.distinct(VendaEvento.marca)).label('marcas_diferentes'),
        func.count(func.distinct(VendaEvento.vendedor)).label('vendedores_diferentes')
    ).group_by(VendaEvento.cliente_nome).order_by(func.count(func.distinct(VendaEvento.numero_nf)).desc()).limit(20).all()
    
    # 2. CAPILARIDADE POR VENDEDOR (CONSULTOR)
    capilaridade_vendedor = db.session.query(
        VendaEvento.vendedor,
        VendaEvento.equipe,
        func.count(func.distinct(VendaEvento.numero_nf)).label('total_nfs'),
        func.count(func.distinct(VendaEvento.cliente_nome)).label('clientes_atendidos'),
        func.sum(VendaEvento.valor_total).label('valor_total'),
        func.sum(VendaEvento.quantidade).label('quantidade_total'),
        func.count(func.distinct(VendaEvento.marca)).label('marcas_vendidas')
    ).group_by(VendaEvento.vendedor, VendaEvento.equipe).order_by(func.count(func.distinct(VendaEvento.numero_nf)).desc()).limit(20).all()
    
    # 3. CAPILARIDADE POR MARCA
    capilaridade_marca = db.session.query(
        VendaEvento.marca,
        func.count(func.distinct(VendaEvento.numero_nf)).label('total_nfs'),
        func.count(func.distinct(VendaEvento.cliente_nome)).label('clientes_diferentes'),
        func.count(func.distinct(VendaEvento.vendedor)).label('vendedores_diferentes'),
        func.sum(VendaEvento.valor_total).label('valor_total'),
        func.sum(VendaEvento.quantidade).label('quantidade_total'),
        func.count(VendaEvento.id).label('total_itens')
    ).group_by(VendaEvento.marca).order_by(func.count(func.distinct(VendaEvento.numero_nf)).desc()).limit(20).all()
    
    # 4. ESTATÍSTICAS GERAIS
    total_nfs = db.session.query(func.count(func.distinct(VendaEvento.numero_nf))).scalar() or 0
    total_clientes = db.session.query(func.count(func.distinct(VendaEvento.cliente_nome))).scalar() or 0
    total_vendedores = db.session.query(func.count(func.distinct(VendaEvento.vendedor))).scalar() or 0
    total_marcas = db.session.query(func.count(func.distinct(VendaEvento.marca))).scalar() or 0
    
    # CORREÇÃO: Média de itens por NF usando subquery
    # Subquery para contar itens por NF
    subquery_itens = db.session.query(
        VendaEvento.numero_nf,
        func.count(VendaEvento.id).label('total_itens')
    ).group_by(VendaEvento.numero_nf).subquery()
    
    # Subquery para somar valor por NF
    subquery_valor = db.session.query(
        VendaEvento.numero_nf,
        func.sum(VendaEvento.valor_total).label('valor_total')
    ).group_by(VendaEvento.numero_nf).subquery()
    
    # Média de itens por NF
    media_itens_por_nf = db.session.query(
        func.avg(subquery_itens.c.total_itens)
    ).scalar() or 0
    
    # Média de valor por NF
    media_valor_por_nf = db.session.query(
        func.avg(subquery_valor.c.valor_total)
    ).scalar() or 0
    
    return jsonify({
        'estatisticas_gerais': {
            'total_nfs': int(total_nfs),
            'total_clientes': int(total_clientes),
            'total_vendedores': int(total_vendedores),
            'total_marcas': int(total_marcas),
            'media_itens_por_nf': float(media_itens_por_nf),
            'media_valor_por_nf': float(media_valor_por_nf)
        },
        'capilaridade_cliente': [
            {
                'cliente': cliente,
                'total_nfs': int(total_nfs),
                'total_itens': int(total_itens),
                'valor_total': float(valor_total),
                'quantidade_total': int(quantidade_total),
                'marcas_diferentes': int(marcas_diferentes),
                'vendedores_diferentes': int(vendedores_diferentes),
                'media_itens_por_nf': float(total_itens / total_nfs) if total_nfs > 0 else 0,
                'media_valor_por_nf': float(valor_total / total_nfs) if total_nfs > 0 else 0
            }
            for cliente, total_nfs, total_itens, valor_total, quantidade_total, marcas_diferentes, vendedores_diferentes in capilaridade_cliente
        ],
        'capilaridade_vendedor': [
            {
                'vendedor': vendedor,
                'equipe': equipe,
                'total_nfs': int(total_nfs),
                'clientes_atendidos': int(clientes_atendidos),
                'valor_total': float(valor_total),
                'quantidade_total': int(quantidade_total),
                'marcas_vendidas': int(marcas_vendidas),
                'media_nfs_por_cliente': float(total_nfs / clientes_atendidos) if clientes_atendidos > 0 else 0,
                'media_valor_por_nf': float(valor_total / total_nfs) if total_nfs > 0 else 0
            }
            for vendedor, equipe, total_nfs, clientes_atendidos, valor_total, quantidade_total, marcas_vendidas in capilaridade_vendedor
        ],
        'capilaridade_marca': [
            {
                'marca': marca,
                'total_nfs': int(total_nfs),
                'clientes_diferentes': int(clientes_diferentes),
                'vendedores_diferentes': int(vendedores_diferentes),
                'valor_total': float(valor_total),
                'quantidade_total': int(quantidade_total),
                'total_itens': int(total_itens),
                'media_itens_por_nf': float(total_itens / total_nfs) if total_nfs > 0 else 0,
                'media_valor_por_nf': float(valor_total / total_nfs) if total_nfs > 0 else 0
            }
            for marca, total_nfs, clientes_diferentes, vendedores_diferentes, valor_total, quantidade_total, total_itens in capilaridade_marca
        ]
    })

@app.route('/api/analise-detalhada-marca', methods=['POST'])
@login_required
def api_analise_detalhada_marca():
    """Análise detalhada por marca específica COM FILTROS e cálculo CROS"""
    data = request.get_json()
    
    if not data.get('marca_especifica'):
        return jsonify({'error': 'Marca não especificada'}), 400
    
    marca = data['marca_especifica']
    
    # Construir query base para a marca específica
    query_marca = VendaEvento.query.filter(VendaEvento.marca == marca)
    
    # Construir query base para totais (com filtros, exceto marca)
    query_totais = VendaEvento.query
    
    # Aplicar filtros às duas queries
    filtros_comuns = ['cliente', 'vendedor', 'equipe', 'familia', 'data_inicio', 'data_fim']
    
    for filtro in filtros_comuns:
        if data.get(filtro) and data[filtro] != 'todos':
            if filtro == 'data_inicio':
                try:
                    data_inicio = datetime.strptime(data['data_inicio'], '%Y-%m-%d').date()
                    query_marca = query_marca.filter(VendaEvento.data_emissao >= data_inicio)
                    query_totais = query_totais.filter(VendaEvento.data_emissao >= data_inicio)
                except:
                    pass
            elif filtro == 'data_fim':
                try:
                    data_fim = datetime.strptime(data['data_fim'], '%Y-%m-%d').date()
                    query_marca = query_marca.filter(VendaEvento.data_emissao <= data_fim)
                    query_totais = query_totais.filter(VendaEvento.data_emissao <= data_fim)
                except:
                    pass
            else:
                campo = getattr(VendaEvento, {
                    'cliente': 'cliente_nome',
                    'vendedor': 'vendedor',
                    'equipe': 'equipe',
                    'familia': 'familia'
                }[filtro])
                
                query_marca = query_marca.filter(campo == data[filtro])
                query_totais = query_totais.filter(campo == data[filtro])
    
    from sqlalchemy import func
    
    # 1. DADOS GERAIS DA MARCA
    dados_marca = query_marca.with_entities(
        func.count(func.distinct(VendaEvento.numero_nf)).label('total_nfs'),
        func.sum(VendaEvento.valor_total).label('total_valor'),
        func.sum(VendaEvento.quantidade).label('total_quantidade'),
        func.count(VendaEvento.id).label('total_itens')
    ).first()
    
    # 2. TICKET MÉDIO POR CLIENTE
    ticket_medio_query = query_marca.with_entities(
        VendaEvento.cliente_nome,
        func.sum(VendaEvento.valor_total).label('valor_total')
    ).group_by(VendaEvento.cliente_nome).subquery()
    
    ticket_medio_cliente = db.session.query(
        func.avg(ticket_medio_query.c.valor_total)
    ).scalar() or 0
    
    # 3. CLIENTES COMPRADORES DA MARCA
    clientes_compradores_marca = query_marca.with_entities(
        func.count(func.distinct(VendaEvento.cliente_nome))
    ).scalar() or 0
    
    # 4. TOTAL DE CLIENTES (todas marcas, com filtros aplicados)
    total_clientes_geral = query_totais.with_entities(
        func.count(func.distinct(VendaEvento.cliente_nome))
    ).scalar() or 0
    
    # 5. CÁLCULO VENDAS CROS (NFs com múltiplas marcas)
    # Primeiro, identificar NFs que têm a marca específica
    nfs_da_marca = query_marca.with_entities(
        func.distinct(VendaEvento.numero_nf)
    ).subquery()
    
    # Agora verificar quais dessas NFs têm mais de uma marca
    nfs_multimarca = db.session.query(
        VendaEvento.numero_nf,
        func.count(func.distinct(VendaEvento.marca)).label('qtd_marcas'),
        func.sum(VendaEvento.valor_total).label('valor_total_nf')
    ).filter(VendaEvento.numero_nf.in_(
        db.session.query(nfs_da_marca)
    )).group_by(VendaEvento.numero_nf).having(
        func.count(func.distinct(VendaEvento.marca)) > 1
    ).subquery()
    
    # Calcular valor CROS (só o valor da marca específica nessas NFs mistas)
    vendas_cros = db.session.query(
        func.sum(VendaEvento.valor_total).label('valor_cros'),
        func.count(func.distinct(VendaEvento.numero_nf)).label('nfs_cros')
    ).filter(
        VendaEvento.marca == marca,
        VendaEvento.numero_nf.in_(
            db.session.query(nfs_multimarca.c.numero_nf)
        )
    ).first()
    
    # 6. TOP PRODUTOS DA MARCA (por VALOR)
    top_produtos = query_marca.with_entities(
        VendaEvento.descricao_produto,
        VendaEvento.familia,
        func.sum(VendaEvento.valor_total).label('total_valor'),
        func.sum(VendaEvento.quantidade).label('total_quantidade'),
        func.count(func.distinct(VendaEvento.numero_nf)).label('total_nfs'),
        func.avg(VendaEvento.valor_total / VendaEvento.quantidade).label('valor_unitario_medio')
    ).group_by(VendaEvento.descricao_produto, VendaEvento.familia).order_by(
        func.sum(VendaEvento.valor_total).desc()
    ).limit(15).all()
    
    # 7. TOTAL GERAL (para cálculo de porcentagem)
    total_geral = query_totais.with_entities(
        func.sum(VendaEvento.valor_total)
    ).scalar() or 0
    
    # 8. TOP CLIENTES DA MARCA (opcional - se quiser mostrar na tabela)
    top_clientes = query_marca.with_entities(
        VendaEvento.cliente_nome,
        func.sum(VendaEvento.valor_total).label('total_valor'),
        func.count(func.distinct(VendaEvento.numero_nf)).label('total_nfs')
    ).group_by(VendaEvento.cliente_nome).order_by(
        func.sum(VendaEvento.valor_total).desc()
    ).limit(10).all()
    
    return jsonify({
        'dados_gerais': {
            'total_nfs': int(dados_marca[0]) if dados_marca[0] else 0,
            'total_valor': float(dados_marca[1]) if dados_marca[1] else 0,
            'total_quantidade': int(dados_marca[2]) if dados_marca[2] else 0,
            'total_itens': int(dados_marca[3]) if dados_marca[3] else 0,
            'ticket_medio_cliente': float(ticket_medio_cliente),
            'clientes_compradores': int(clientes_compradores_marca),
            'vendas_cros': {
                'valor_total': float(vendas_cros[0]) if vendas_cros[0] else 0,
                'total_nfs': int(vendas_cros[1]) if vendas_cros[1] else 0,
                'percentual_faturamento': float((vendas_cros[0] / dados_marca[1] * 100) if dados_marca[1] and dados_marca[1] > 0 else 0)
            },
            'percentual_total': float((dados_marca[1] / total_geral * 100) if total_geral > 0 else 0)
        },
        'estatisticas_clientes': {
            'total_clientes': int(total_clientes_geral)
        },
        'clientes_compradores': {
            'total_clientes': int(total_clientes_geral),
            'compradores_marca': int(clientes_compradores_marca),
            'percentual_penetracao': float((clientes_compradores_marca / total_clientes_geral * 100) if total_clientes_geral > 0 else 0)
        },
        'top_produtos': [
            {
                'produto': produto or 'Desconhecido',
                'descricao': produto or 'Desconhecido',  # Mantido para compatibilidade
                'familia': familia or 'N/A',
                'valor_total': float(valor),
                'quantidade': int(quantidade),
                'nfs': int(nfs),
                'valor_unitario_medio': float(valor_unitario) if valor_unitario else 0,
                'percentual_marca': float((valor / dados_marca[1] * 100) if dados_marca[1] and dados_marca[1] > 0 else 0)
            }
            for produto, familia, valor, quantidade, nfs, valor_unitario in top_produtos
        ],
        'top_clientes': [
            {
                'cliente': cliente,
                'valor': float(valor),
                'nfs': int(nfs),
                'ticket_medio': float(valor / nfs if nfs > 0 else 0)
            }
            for cliente, valor, nfs in top_clientes
        ]
    })

# Adicione esta rota para métricas com filtros
@app.route('/api/metricas-vendas-filtradas', methods=['POST'])
@login_required
def api_metricas_vendas_filtradas():
    """API para métricas consolidadas com filtros - ATUALIZADA COM PORCENTAGENS"""
    data = request.get_json()
    
    # Construir query base
    query = VendaEvento.query
    
    # Aplicar filtros
    if data.get('cliente') and data['cliente'] != 'todos':
        query = query.filter(VendaEvento.cliente_nome == data['cliente'])
    
    if data.get('vendedor') and data['vendedor'] != 'todos':
        query = query.filter(VendaEvento.vendedor == data['vendedor'])
    
    if data.get('equipe') and data['equipe'] != 'todos':
        query = query.filter(VendaEvento.equipe == data['equipe'])
    
    if data.get('marca') and data['marca'] != 'todos':
        query = query.filter(VendaEvento.marca == data['marca'])
    
    if data.get('familia') and data['familia'] != 'todos':
        query = query.filter(VendaEvento.familia == data['familia'])
    
    if data.get('data_inicio'):
        try:
            data_inicio = datetime.strptime(data['data_inicio'], '%Y-%m-%d').date()
            query = query.filter(VendaEvento.data_emissao >= data_inicio)
        except:
            pass
    
    if data.get('data_fim'):
        try:
            data_fim = datetime.strptime(data['data_fim'], '%Y-%m-%d').date()
            query = query.filter(VendaEvento.data_emissao <= data_fim)
        except:
            pass
    
    from sqlalchemy import func
    
    # TOTAIS GERAIS
    total_nfs = query.with_entities(func.count(func.distinct(VendaEvento.numero_nf))).scalar() or 0
    total_vendas_itens = query.count()
    total_valor = query.with_entities(func.sum(VendaEvento.valor_total)).scalar() or 0
    total_quantidade = query.with_entities(func.sum(VendaEvento.quantidade)).scalar() or 0
    
    # TOTAL DE CLIENTES ÚNICOS (para cálculo de penetração)
    total_clientes = query.with_entities(func.count(func.distinct(VendaEvento.cliente_nome))).scalar() or 0
    
    # Calcular valores médios
    valor_medio_item = total_valor / total_quantidade if total_quantidade > 0 else 0
    
    # Média de itens por NF
    if total_nfs > 0:
        subquery_itens = query.with_entities(
            VendaEvento.numero_nf,
            func.count(VendaEvento.id).label('total_itens')
        ).group_by(VendaEvento.numero_nf).subquery()
        
        media_itens_por_nf = db.session.query(
            func.avg(subquery_itens.c.total_itens)
        ).scalar() or 0
    else:
        media_itens_por_nf = 0
    
    # Média de valor por NF
    if total_nfs > 0:
        subquery_valor = query.with_entities(
            VendaEvento.numero_nf,
            func.sum(VendaEvento.valor_total).label('valor_total')
        ).group_by(VendaEvento.numero_nf).subquery()
        
        media_valor_por_nf = db.session.query(
            func.avg(subquery_valor.c.valor_total)
        ).scalar() or 0
    else:
        media_valor_por_nf = 0
    
    # Cliente que mais comprou (por valor) COM FILTROS
    cliente_maior_valor = query.with_entities(
        VendaEvento.cliente_nome,
        func.sum(VendaEvento.valor_total).label('total_valor'),
        func.count(func.distinct(VendaEvento.numero_nf)).label('total_nfs')
    ).group_by(VendaEvento.cliente_nome).order_by(func.sum(VendaEvento.valor_total).desc()).first()
    
    # Vendedor com maior valor COM FILTROS
    vendedor_maior_valor = query.with_entities(
        VendaEvento.vendedor,
        func.sum(VendaEvento.valor_total).label('total_valor'),
        func.count(func.distinct(VendaEvento.numero_nf)).label('total_nfs')
    ).group_by(VendaEvento.vendedor).order_by(func.sum(VendaEvento.valor_total).desc()).first()
    
    # Marcas por quantidade de NFs COM FILTROS E PORCENTAGENS
    marcas_por_nf = query.with_entities(
        VendaEvento.marca,
        func.count(func.distinct(VendaEvento.numero_nf)).label('total_nfs'),
        func.sum(VendaEvento.valor_total).label('total_valor'),
        func.sum(VendaEvento.quantidade).label('total_quantidade')
    ).group_by(VendaEvento.marca).order_by(func.sum(VendaEvento.valor_total).desc()).limit(10).all()
    
    # Calcular porcentagens para cada marca
    marcas_com_porcentagem = []
    for marca, total_nfs_marca, valor_marca, quantidade_marca in marcas_por_nf:
        percentual_valor = (valor_marca / total_valor * 100) if total_valor > 0 else 0
        marcas_com_porcentagem.append({
            'marca': marca or 'Desconhecida',
            'nfs': int(total_nfs_marca),
            'valor': float(valor_marca),
            'quantidade': int(quantidade_marca),
            'percentual_valor': float(percentual_valor)
        })
    
    return jsonify({
        'totais_por_nf': {
            'total_nfs': int(total_nfs),
            'total_itens': int(total_vendas_itens),
            'total_valor': float(total_valor),
            'total_quantidade': int(total_quantidade),
            'media_itens_por_nf': float(media_itens_por_nf),
            'media_valor_por_nf': float(media_valor_por_nf),
            'media_valor_por_item': float(valor_medio_item)
        },
        'estatisticas_clientes': {
            'total_clientes': int(total_clientes)
        },
        'clientes': {
            'maior_valor': {
                'nome': cliente_maior_valor[0] if cliente_maior_valor else None,
                'valor': float(cliente_maior_valor[1]) if cliente_maior_valor else 0,
                'nfs': cliente_maior_valor[2] if cliente_maior_valor else 0
            }
        },
        'vendedores': {
            'maior_valor': {
                'nome': vendedor_maior_valor[0] if vendedor_maior_valor else None,
                'valor': float(vendedor_maior_valor[1]) if vendedor_maior_valor else 0,
                'nfs': vendedor_maior_valor[2] if vendedor_maior_valor else 0
            }
        },
        'marcas_por_nf': marcas_com_porcentagem
    })       
    
@app.route('/api/capilaridade-vendas-filtradas', methods=['POST'])
@login_required
def api_capilaridade_vendas_filtradas():
    """API para análise de capilaridade com filtros - SIMPLIFICADA (sem top 10 marcas)"""
    data = request.get_json()
    
    # Construir query base
    query = VendaEvento.query
    
    # Aplicar filtros
    if data.get('cliente') and data['cliente'] != 'todos':
        query = query.filter(VendaEvento.cliente_nome == data['cliente'])
    
    if data.get('vendedor') and data['vendedor'] != 'todos':
        query = query.filter(VendaEvento.vendedor == data['vendedor'])
    
    if data.get('equipe') and data['equipe'] != 'todos':
        query = query.filter(VendaEvento.equipe == data['equipe'])
    
    if data.get('marca') and data['marca'] != 'todos':
        query = query.filter(VendaEvento.marca == data['marca'])
    
    if data.get('familia') and data['familia'] != 'todos':
        query = query.filter(VendaEvento.familia == data['familia'])
    
    if data.get('data_inicio'):
        try:
            data_inicio = datetime.strptime(data['data_inicio'], '%Y-%m-%d').date()
            query = query.filter(VendaEvento.data_emissao >= data_inicio)
        except:
            pass
    
    if data.get('data_fim'):
        try:
            data_fim = datetime.strptime(data['data_fim'], '%Y-%m-%d').date()
            query = query.filter(VendaEvento.data_emissao <= data_fim)
        except:
            pass
    
    from sqlalchemy import func
    
    # ESTATÍSTICAS GERAIS COM FILTROS
    total_nfs = query.with_entities(func.count(func.distinct(VendaEvento.numero_nf))).scalar() or 0
    total_clientes = query.with_entities(func.count(func.distinct(VendaEvento.cliente_nome))).scalar() or 0
    total_vendedores = query.with_entities(func.count(func.distinct(VendaEvento.vendedor))).scalar() or 0
    total_marcas = query.with_entities(func.count(func.distinct(VendaEvento.marca))).scalar() or 0
    
    # Média de itens por NF
    if total_nfs > 0:
        subquery_itens = query.with_entities(
            VendaEvento.numero_nf,
            func.count(VendaEvento.id).label('total_itens')
        ).group_by(VendaEvento.numero_nf).subquery()
        
        media_itens_por_nf = db.session.query(
            func.avg(subquery_itens.c.total_itens)
        ).scalar() or 0
    else:
        media_itens_por_nf = 0
    
    # Média de valor por NF
    if total_nfs > 0:
        subquery_valor = query.with_entities(
            VendaEvento.numero_nf,
            func.sum(VendaEvento.valor_total).label('valor_total')
        ).group_by(VendaEvento.numero_nf).subquery()
        
        media_valor_por_nf = db.session.query(
            func.avg(subquery_valor.c.valor_total)
        ).scalar() or 0
    else:
        media_valor_por_nf = 0
    
    return jsonify({
        'estatisticas_gerais': {
            'total_nfs': int(total_nfs),
            'total_clientes': int(total_clientes),
            'total_vendedores': int(total_vendedores),
            'total_marcas': int(total_marcas),
            'media_itens_por_nf': float(media_itens_por_nf),
            'media_valor_por_nf': float(media_valor_por_nf)
        }
        # REMOVIDO: capilaridade_marca (não é mais necessário)
    })

@app.route('/api/analise-marca-filtrada/<path:marca>', methods=['POST'])
@login_required
def api_analise_marca_filtrada(marca):
    """Análise detalhada por marca com filtros"""
    try:
        # Decodificar marca (pode ter caracteres especiais)
        from urllib.parse import unquote
        marca = unquote(marca)
        
        data = request.get_json()
        
        # Construir query base
        query = VendaEvento.query.filter(VendaEvento.marca == marca)
        
        # Aplicar filtros
        if data.get('cliente') and data['cliente'] != 'todos':
            query = query.filter(VendaEvento.cliente_nome == data['cliente'])
        
        if data.get('vendedor') and data['vendedor'] != 'todos':
            query = query.filter(VendaEvento.vendedor == data['vendedor'])
        
        if data.get('equipe') and data['equipe'] != 'todos':
            query = query.filter(VendaEvento.equipe == data['equipe'])
        
        if data.get('familia') and data['familia'] != 'todos':
            query = query.filter(VendaEvento.familia == data['familia'])
        
        if data.get('data_inicio'):
            try:
                data_inicio = datetime.strptime(data['data_inicio'], '%Y-%m-%d').date()
                query = query.filter(VendaEvento.data_emissao >= data_inicio)
            except:
                pass
        
        if data.get('data_fim'):
            try:
                data_fim = datetime.strptime(data['data_fim'], '%Y-%m-%d').date()
                query = query.filter(VendaEvento.data_emissao <= data_fim)
            except:
                pass
        
        from sqlalchemy import func
        
        # Dados da marca
        dados_marca = query.with_entities(
            func.count(func.distinct(VendaEvento.numero_nf)).label('total_nfs'),
            func.sum(VendaEvento.valor_total).label('total_valor'),
            func.sum(VendaEvento.quantidade).label('total_quantidade'),
            func.count(VendaEvento.id).label('total_itens'),
            func.avg(VendaEvento.valor_total / VendaEvento.quantidade).label('valor_medio_unitario')
        ).first()
        
        # Top clientes da marca (por NFs)
        top_clientes = query.with_entities(
            VendaEvento.cliente_nome,
            func.count(func.distinct(VendaEvento.numero_nf)).label('total_nfs'),
            func.sum(VendaEvento.valor_total).label('total_valor'),
            func.sum(VendaEvento.quantidade).label('total_quantidade')
        ).group_by(VendaEvento.cliente_nome).order_by(func.count(func.distinct(VendaEvento.numero_nf)).desc()).limit(10).all()
        
        # Top vendedores da marca (por NFs)
        top_vendedores = query.with_entities(
            VendaEvento.vendedor,
            func.count(func.distinct(VendaEvento.numero_nf)).label('total_nfs'),
            func.sum(VendaEvento.valor_total).label('total_valor'),
            func.sum(VendaEvento.quantidade).label('total_quantidade'),
            func.count(func.distinct(VendaEvento.cliente_nome)).label('clientes_atendidos')
        ).group_by(VendaEvento.vendedor).order_by(func.count(func.distinct(VendaEvento.numero_nf)).desc()).limit(10).all()
        
        # Top produtos da marca
        top_produtos = query.with_entities(
            VendaEvento.descricao_produto,
            VendaEvento.familia,
            func.count(func.distinct(VendaEvento.numero_nf)).label('total_nfs'),
            func.sum(VendaEvento.valor_total).label('total_valor'),
            func.sum(VendaEvento.quantidade).label('total_quantidade'),
            func.avg(VendaEvento.valor_total / VendaEvento.quantidade).label('valor_unitario_medio')
        ).group_by(VendaEvento.descricao_produto, VendaEvento.familia).order_by(func.count(func.distinct(VendaEvento.numero_nf)).desc()).limit(15).all()
        
        # Distribuição por família
        distribuicao_familia = query.with_entities(
            VendaEvento.familia,
            func.count(func.distinct(VendaEvento.numero_nf)).label('total_nfs'),
            func.sum(VendaEvento.valor_total).label('total_valor'),
            func.sum(VendaEvento.quantidade).label('total_quantidade'),
            func.count(VendaEvento.id).label('total_itens')
        ).group_by(VendaEvento.familia).all()
        
        return jsonify({
            'marca': marca,
            'dados_gerais': {
                'total_nfs': int(dados_marca[0]) if dados_marca[0] else 0,
                'total_valor': float(dados_marca[1]) if dados_marca[1] else 0,
                'total_quantidade': int(dados_marca[2]) if dados_marca[2] else 0,
                'total_itens': int(dados_marca[3]) if dados_marca[3] else 0,
                'valor_medio_unitario': float(dados_marca[4]) if dados_marca[4] else 0,
                'media_valor_por_nf': float(dados_marca[1] / dados_marca[0]) if dados_marca[0] and dados_marca[1] and dados_marca[0] > 0 else 0,
                'media_itens_por_nf': float(dados_marca[3] / dados_marca[0]) if dados_marca[0] and dados_marca[3] and dados_marca[0] > 0 else 0
            },
            'top_clientes': [
                {
                    'cliente': cliente or 'Desconhecido',
                    'nfs': int(total_nfs) if total_nfs else 0,
                    'valor': float(valor) if valor else 0,
                    'quantidade': int(quantidade) if quantidade else 0,
                    'media_valor_por_nf': float(valor / total_nfs) if total_nfs and total_nfs > 0 else 0,
                    'ticket_medio': float(valor / total_nfs) if total_nfs and total_nfs > 0 else 0
                }
                for cliente, total_nfs, valor, quantidade in top_clientes
            ],
            'top_vendedores': [
                {
                    'vendedor': vendedor or 'Desconhecido',
                    'nfs': int(total_nfs) if total_nfs else 0,
                    'valor': float(valor) if valor else 0,
                    'quantidade': int(quantidade) if quantidade else 0,
                    'clientes_atendidos': int(clientes_atendidos) if clientes_atendidos else 0,
                    'media_nfs_por_cliente': float(total_nfs / clientes_atendidos) if clientes_atendidos and clientes_atendidos > 0 else 0
                }
                for vendedor, total_nfs, valor, quantidade, clientes_atendidos in top_vendedores
            ],
            'top_produtos': [
                {
                    'produto': produto or 'Desconhecido',
                    'familia': familia or 'N/A',
                    'nfs': int(total_nfs) if total_nfs else 0,
                    'valor': float(valor) if valor else 0,
                    'quantidade': int(quantidade) if quantidade else 0,
                    'valor_unitario_medio': float(valor_unitario_medio) if valor_unitario_medio else 0
                }
                for produto, familia, total_nfs, valor, quantidade, valor_unitario_medio in top_produtos
            ],
            'distribuicao_familia': [
                {
                    'familia': familia or 'Sem Família',
                    'nfs': int(total_nfs) if total_nfs else 0,
                    'valor': float(valor) if valor else 0,
                    'quantidade': int(quantidade) if quantidade else 0,
                    'itens': int(total_itens) if total_itens else 0
                }
                for familia, total_nfs, valor, quantidade, total_itens in distribuicao_familia
            ]
        })
        
    except Exception as e:
        print(f"❌ Erro em api_analise_marca_filtrada: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500
    
# Adicione APENAS esta rota nova ao seu app.py - ela é necessária para os filtros
@app.route('/api/metricas-vendas-filtradas-simples', methods=['POST'])
@login_required
def api_metricas_vendas_filtradas_simples():
    """API simples para métricas com filtros - usa as rotas existentes"""
    try:
        data = request.get_json()
        
        # Primeiro, pega os dados completos da rota existente
        response_data = {}
        
        # Tenta pegar dados da rota /api/metricas-vendas
        try:
            from flask import url_for
            import requests
            import json
            
            # Simula uma chamada interna para a rota existente
            with app.test_request_context():
                # Usa a função existente
                from flask import g
                result = api_metricas_vendas()
                if isinstance(result, tuple):
                    response_data.update(result[0].get_json())
                else:
                    response_data.update(result.get_json())
        except:
            # Se falhar, retorna dados básicos
            response_data = {
                'totais_por_nf': {
                    'total_nfs': 0,
                    'total_valor': 0,
                    'total_itens': 0,
                    'media_valor_por_nf': 0
                },
                'vendedores': {'maior_valor': {'nome': 'N/A', 'valor': 0}},
                'clientes': {'maior_valor': {'nome': 'N/A', 'valor': 0}},
                'marcas_por_nf': []
            }
        
        return jsonify(response_data)
        
    except Exception as e:
        print(f"❌ Erro em api_metricas_vendas_filtradas_simples: {str(e)}")
        return jsonify({'error': str(e)}), 500

# Atualize a rota existente /exportar-vendas-filtradas para aceitar filtros via POST
@app.route('/exportar-vendas-filtradas', methods=['GET', 'POST'])
@login_required
@permissao_required('exportacao')
def exportar_vendas_filtradas():
    """Exportar vendas com filtros aplicados para Excel"""
    # Se for POST, obter filtros do corpo da requisição
    if request.method == 'POST':
        data = request.get_json()
        filtros = data
    else:
        # Se for GET, obter filtros da query string
        filtros = {
            'cliente': request.args.get('cliente', 'todos'),
            'vendedor': request.args.get('vendedor', 'todos'),
            'marca': request.args.get('marca', 'todos'),
            'equipe': request.args.get('equipe', 'todos'),
            'familia': request.args.get('familia', 'todos'),
            'data_inicio': request.args.get('data_inicio'),
            'data_fim': request.args.get('data_fim')
        }
    
    # Construir query com filtros
    query = VendaEvento.query
    
    if filtros.get('cliente') and filtros['cliente'] != 'todos':
        query = query.filter(VendaEvento.cliente_nome == filtros['cliente'])
    
    if filtros.get('vendedor') and filtros['vendedor'] != 'todos':
        query = query.filter(VendaEvento.vendedor == filtros['vendedor'])
    
    if filtros.get('equipe') and filtros['equipe'] != 'todos':
        query = query.filter(VendaEvento.equipe == filtros['equipe'])
    
    if filtros.get('marca') and filtros['marca'] != 'todos':
        query = query.filter(VendaEvento.marca == filtros['marca'])
    
    if filtros.get('familia') and filtros['familia'] != 'todos':
        query = query.filter(VendaEvento.familia == filtros['familia'])
    
    if filtros.get('data_inicio'):
        try:
            data_inicio = datetime.strptime(filtros['data_inicio'], '%Y-%m-%d').date()
            query = query.filter(VendaEvento.data_emissao >= data_inicio)
        except:
            pass
    
    if filtros.get('data_fim'):
        try:
            data_fim = datetime.strptime(filtros['data_fim'], '%Y-%m-%d').date()
            query = query.filter(VendaEvento.data_emissao <= data_fim)
        except:
            pass
    
    # Ordenar por NF e data
    query = query.order_by(VendaEvento.numero_nf, VendaEvento.data_emissao.desc())
    
    # Buscar todas as vendas
    vendas = query.all()
    
    # Preparar dados para exportação
    data = []
    for venda in vendas:
        # Calcular valor unitário
        valor_unitario = venda.valor_total / venda.quantidade if venda.quantidade > 0 else venda.valor_total
        
        data.append({
            'NF': venda.numero_nf,
            'DATA_EMISSAO': venda.data_emissao.strftime('%d/%m/%Y'),
            'CLIENTE_NOME': venda.cliente_nome,
            'VENDEDOR': venda.vendedor,
            'EQUIPE': venda.equipe,
            'DESCRICAO_PRODUTO': venda.descricao_produto,
            'MARCA': venda.marca,
            'FAMILIA': venda.familia or 'N/A',
            'VALOR_TOTAL': float(venda.valor_total),
            'QUANTIDADE': venda.quantidade,
            'VALOR_UNITARIO': float(valor_unitario)
        })
    
    output = export_to_excel(data, 'vendas_filtradas.xlsx', 'Vendas Filtradas')
    
    # Registrar log
    registrar_log('exportacao_vendas_filtradas', 'exportacao', {
        'quantidade_vendas': len(vendas),
        'filtros_aplicados': filtros
    })
    
    return send_file(
        output,
        download_name='vendas_filtradas_ro_experience.xlsx',
        as_attachment=True,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

@app.route('/download-template-vendas')
@login_required
def download_template_vendas():
    """Download do template para importação de vendas - ATUALIZADO COM NF"""
    data = [
        {
            'NF': '123456',
            'DATA_EMISSAO': '15/10/2025',
            'CLIENTE_NOME': 'Empresa ABC Ltda',
            'VENDEDOR': 'João Silva',
            'EQUIPE': 'Equipe Sul',
            'DESCRICAO_PRODUTO': 'Produto X Premium',
            'MARCA': 'Marca A',
            'VALOR_PRODUTOS': '1500.00',  # Valor TOTAL da venda
            'QTD': '2',
            'FAMILIA': 'Premium'
        },
        {
            'NF': '123456',  # MESMA NF - produto diferente
            'DATA_EMISSAO': '15/10/2025',
            'CLIENTE_NOME': 'Empresa ABC Ltda',
            'VENDEDOR': 'João Silva',
            'EQUIPE': 'Equipe Sul',
            'DESCRICAO_PRODUTO': 'Produto Y Standard',
            'MARCA': 'Marca A',
            'VALOR_PRODUTOS': '800.50',
            'QTD': '3',
            'FAMILIA': 'Standard'
        },
        {
            'NF': '789012',
            'DATA_EMISSAO': '15/10/2025',
            'CLIENTE_NOME': 'Comércio XYZ S/A',
            'VENDEDOR': 'Maria Santos',
            'EQUIPE': 'Equipe Norte',
            'DESCRICAO_PRODUTO': 'Produto Z Basic',
            'MARCA': 'Marca B',
            'VALOR_PRODUTOS': '300.00',
            'QTD': '5',
            'FAMILIA': 'Basic'
        }
    ]
    
    output = export_to_excel(data, 'template_vendas_ro_experience.xlsx', 'Template Vendas')
    
    return send_file(
        output,
        download_name='template_vendas_ro_experience.xlsx',
        as_attachment=True,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
@app.route('/exportar-analise-imagem', methods=['POST'])
@login_required
def exportar_analise_imagem():
    """API para exportar análise como imagem com filtros aplicados - ATUALIZADA COM NOVOS DADOS"""
    try:
        data = request.get_json()
        filtros = data.get('filtros', {})
        
        # Construir query com filtros
        query = VendaEvento.query
        
        if filtros.get('cliente') and filtros['cliente'] != 'todos':
            query = query.filter(VendaEvento.cliente_nome == filtros['cliente'])
        
        if filtros.get('vendedor') and filtros['vendedor'] != 'todos':
            query = query.filter(VendaEvento.vendedor == filtros['vendedor'])
        
        if filtros.get('equipe') and filtros['equipe'] != 'todos':
            query = query.filter(VendaEvento.equipe == filtros['equipe'])
        
        if filtros.get('marca') and filtros['marca'] != 'todos':
            query = query.filter(VendaEvento.marca == filtros['marca'])
        
        if filtros.get('familia') and filtros['familia'] != 'todos':
            query = query.filter(VendaEvento.familia == filtros['familia'])
        
        if filtros.get('data_inicio'):
            try:
                data_inicio = datetime.strptime(filtros['data_inicio'], '%Y-%m-%d').date()
                query = query.filter(VendaEvento.data_emissao >= data_inicio)
            except:
                pass
        
        if filtros.get('data_fim'):
            try:
                data_fim = datetime.strptime(filtros['data_fim'], '%Y-%m-%d').date()
                query = query.filter(VendaEvento.data_emissao <= data_fim)
            except:
                pass
        
        from sqlalchemy import func
        
        # 1. Métricas Consolidadas (mantido)
        total_nfs = query.with_entities(func.count(func.distinct(VendaEvento.numero_nf))).scalar() or 0
        total_valor = query.with_entities(func.sum(VendaEvento.valor_total)).scalar() or 0
        total_itens = query.with_entities(func.sum(VendaEvento.quantidade)).scalar() or 0
        media_valor_por_nf = total_valor / total_nfs if total_nfs > 0 else 0
        
        # Consultor top
        consultor_top_result = query.with_entities(
            VendaEvento.vendedor,
            func.sum(VendaEvento.valor_total).label('total_valor')
        ).group_by(VendaEvento.vendedor).order_by(func.sum(VendaEvento.valor_total).desc()).first()
        
        consultor_top = {
            'nome': consultor_top_result[0] if consultor_top_result and consultor_top_result[0] else 'N/A',
            'valor': float(consultor_top_result[1]) if consultor_top_result and consultor_top_result[1] else 0
        }
        
        # Cliente top
        cliente_top_result = query.with_entities(
            VendaEvento.cliente_nome,
            func.sum(VendaEvento.valor_total).label('total_valor')
        ).group_by(VendaEvento.cliente_nome).order_by(func.sum(VendaEvento.valor_total).desc()).first()
        
        cliente_top = {
            'nome': cliente_top_result[0] if cliente_top_result and cliente_top_result[0] else 'N/A',
            'valor': float(cliente_top_result[1]) if cliente_top_result and cliente_top_result[1] else 0
        }
        
        # 2. Distribuição por Marca COM PORCENTAGENS (ATUALIZADO)
        marcas_result = query.with_entities(
            VendaEvento.marca,
            func.sum(VendaEvento.valor_total).label('valor_total'),
            func.count(func.distinct(VendaEvento.numero_nf)).label('total_nfs')
        ).group_by(VendaEvento.marca).order_by(func.sum(VendaEvento.valor_total).desc()).all()
        
        marcas = []
        for marca in marcas_result:
            percentual = (marca[1] / total_valor * 100) if total_valor > 0 else 0
            marcas.append({
                'marca': marca[0] if marca[0] else 'Desconhecida',
                'valor': float(marca[1]) if marca[1] else 0,
                'nfs': int(marca[2]) if marca[2] else 0,
                'percentual': float(percentual)  # NOVO: porcentagem incluída
            })
        
        # 3. Capilaridade Geral (mantido)
        total_clientes = query.with_entities(func.count(func.distinct(VendaEvento.cliente_nome))).scalar() or 0
        total_vendedores = query.with_entities(func.count(func.distinct(VendaEvento.vendedor))).scalar() or 0
        total_marcas_count = query.with_entities(func.count(func.distinct(VendaEvento.marca))).scalar() or 0
        
        # Média de itens por NF
        if total_nfs > 0:
            subquery_itens = query.with_entities(
                VendaEvento.numero_nf,
                func.count(VendaEvento.id).label('total_itens')
            ).group_by(VendaEvento.numero_nf).subquery()
            
            media_itens_por_nf = db.session.query(
                func.avg(subquery_itens.c.total_itens)
            ).scalar() or 0
        else:
            media_itens_por_nf = 0
        
        # Média de valor por NF (já calculada acima)
        media_valor_por_nf = media_valor_por_nf
        
        # 4. ANÁLISE POR MARCA COM NOVOS DADOS (ATUALIZADO)
        marca_analise = None
        if filtros.get('marca') and filtros['marca'] != 'todos':
            marca_nome = filtros['marca']
        elif marcas:
            marca_nome = marcas[0]['marca']  # Primeira marca do resultado
        else:
            marca_nome = None
        
        if marca_nome:
            marca_query = query.filter(VendaEvento.marca == marca_nome)
            
            # Dados gerais da marca
            marca_dados_result = marca_query.with_entities(
                func.count(func.distinct(VendaEvento.numero_nf)).label('total_nfs'),
                func.sum(VendaEvento.valor_total).label('total_valor'),
                func.sum(VendaEvento.quantidade).label('total_quantidade')
            ).first()
            
            marca_dados = {
                'total_nfs': int(marca_dados_result[0]) if marca_dados_result and marca_dados_result[0] else 0,
                'total_valor': float(marca_dados_result[1]) if marca_dados_result and marca_dados_result[1] else 0,
                'total_quantidade': int(marca_dados_result[2]) if marca_dados_result and marca_dados_result[2] else 0
            }
            
            # 4.1 TICKET MÉDIO POR CLIENTE (NOVO)
            ticket_query = marca_query.with_entities(
                VendaEvento.cliente_nome,
                func.sum(VendaEvento.valor_total).label('valor_total')
            ).group_by(VendaEvento.cliente_nome).subquery()
            
            ticket_medio_cliente = db.session.query(
                func.avg(ticket_query.c.valor_total)
            ).scalar() or 0
            
            # 4.2 CLIENTES VS COMPRADORES (NOVO)
            compradores_marca = marca_query.with_entities(
                func.count(func.distinct(VendaEvento.cliente_nome))
            ).scalar() or 0
            
            percentual_penetracao = (compradores_marca / total_clientes * 100) if total_clientes > 0 else 0
            
            # 4.3 VENDAS CROS (NOVO) - NFs com múltiplas marcas
            # Primeiro, identificar NFs que têm a marca específica
            nfs_da_marca = marca_query.with_entities(
                func.distinct(VendaEvento.numero_nf)
            ).subquery()
            
            # Verificar quais dessas NFs têm mais de uma marca
            nfs_multimarca = db.session.query(
                VendaEvento.numero_nf,
                func.count(func.distinct(VendaEvento.marca)).label('qtd_marcas'),
                func.sum(VendaEvento.valor_total).label('valor_total_nf')
            ).filter(VendaEvento.numero_nf.in_(
                db.session.query(nfs_da_marca)
            )).group_by(VendaEvento.numero_nf).having(
                func.count(func.distinct(VendaEvento.marca)) > 1
            ).subquery()
            
            # Calcular valor CROS (só o valor da marca específica nessas NFs mistas)
            vendas_cros_result = db.session.query(
                func.sum(VendaEvento.valor_total).label('valor_cros'),
                func.count(func.distinct(VendaEvento.numero_nf)).label('nfs_cros')
            ).filter(
                VendaEvento.marca == marca_nome,
                VendaEvento.numero_nf.in_(
                    db.session.query(nfs_multimarca.c.numero_nf)
                )
            ).first()
            
            vendas_cros = {
                'valor_total': float(vendas_cros_result[0]) if vendas_cros_result and vendas_cros_result[0] else 0,
                'total_nfs': int(vendas_cros_result[1]) if vendas_cros_result and vendas_cros_result[1] else 0,
                'percentual_faturamento': float((vendas_cros_result[0] / marca_dados['total_valor'] * 100) 
                                                if marca_dados['total_valor'] and marca_dados['total_valor'] > 0 else 0)
            }
            
            # 4.4 TOP 10 PRODUTOS POR VALOR (NOVO - ordenado por valor)
            top_produtos_result = marca_query.with_entities(
                VendaEvento.descricao_produto,
                func.sum(VendaEvento.valor_total).label('total_valor'),
                func.sum(VendaEvento.quantidade).label('total_quantidade'),
                func.count(func.distinct(VendaEvento.numero_nf)).label('total_nfs')
            ).group_by(VendaEvento.descricao_produto).order_by(
                func.sum(VendaEvento.valor_total).desc()  # Ordenado por VALOR
            ).limit(10).all()
            
            top_produtos = []
            for produto in top_produtos_result:
                percentual_produto = (produto[1] / marca_dados['total_valor'] * 100) if marca_dados['total_valor'] > 0 else 0
                top_produtos.append({
                    'produto': produto[0] if produto[0] else 'Desconhecido',
                    'descricao': produto[0] if produto[0] else 'Desconhecido',
                    'valor': float(produto[1]) if produto[1] else 0,
                    'quantidade': int(produto[2]) if produto[2] else 0,
                    'nfs': int(produto[3]) if produto[3] else 0,
                    'percentual_marca': float(percentual_produto)
                })
            
            # 4.5 TOP CLIENTES DA MARCA (opcional)
            top_clientes_result = marca_query.with_entities(
                VendaEvento.cliente_nome,
                func.count(func.distinct(VendaEvento.numero_nf)).label('total_nfs'),
                func.sum(VendaEvento.valor_total).label('total_valor')
            ).group_by(VendaEvento.cliente_nome).order_by(
                func.sum(VendaEvento.valor_total).desc()
            ).limit(10).all()
            
            top_clientes = []
            for cliente in top_clientes_result:
                top_clientes.append({
                    'cliente': cliente[0] if cliente[0] else 'Desconhecido',
                    'nfs': int(cliente[1]) if cliente[1] else 0,
                    'valor': float(cliente[2]) if cliente[2] else 0
                })
            
            marca_analise = {
                'nome': marca_nome,
                'dados': marca_dados,
                'ticket_medio_cliente': float(ticket_medio_cliente),  # NOVO
                'clientes_compradores': int(compradores_marca),      # NOVO
                'total_clientes_geral': int(total_clientes),         # NOVO
                'percentual_penetracao': float(percentual_penetracao), # NOVO
                'vendas_cros': vendas_cros,                          # NOVO
                'top_produtos': top_produtos,                        # ATUALIZADO
                'top_clientes': top_clientes
            }
        
        # 5. PREPARAR DADOS PARA RETORNO COM NOVA ESTRUTURA
        dados_exportacao = {
            'filtros_aplicados': {
                'cliente': filtros.get('cliente', 'todos'),
                'vendedor': filtros.get('vendedor', 'todos'),
                'marca': filtros.get('marca', 'todos'),
                'equipe': filtros.get('equipe', 'todos'),
                'familia': filtros.get('familia', 'todos'),
                'data_inicio': filtros.get('data_inicio', ''),
                'data_fim': filtros.get('data_fim', ''),
                'data_geracao': agora().strftime('%d/%m/%Y %H:%M')
            },
            'metricas_consolidadas': {
                'total_nfs': int(total_nfs),
                'total_valor': float(total_valor),
                'total_itens': int(total_itens),
                'ticket_medio': float(media_valor_por_nf),
                'consultor_top': consultor_top,
                'cliente_top': cliente_top
            },
            'distribuicao_marca': marcas,  # AGORA COM PORCENTAGENS
            'capilaridade_geral': {
                'total_clientes': int(total_clientes),
                'total_vendedores': int(total_vendedores),
                'total_marcas': int(total_marcas_count),
                'media_itens_por_nf': float(media_itens_por_nf),
                'media_valor_por_nf': float(media_valor_por_nf)
            }
        }
        
        # Adicionar análise da marca com NOVOS DADOS
        if marca_analise:
            dados_exportacao['analise_marca'] = marca_analise
        
        # Registrar log
        registrar_log('exportacao_analise_imagem', 'analise_vendas', {
            'filtros': filtros,
            'metricas_geradas': len(dados_exportacao['distribuicao_marca']),
            'inclui_analise_marca': marca_analise is not None
        })
        
        return jsonify({
            'success': True,
            'dados': dados_exportacao
        })
        
    except Exception as e:
        print(f"❌ Erro ao exportar análise para imagem: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': f'Erro: {str(e)}'})
           
if __name__ == '__main__':
    with app.app_context():
        try:
            # 1. PASSO CRÍTICO: Cria o banco de dados alvo se não existir
            criar_banco_se_nao_existir(app) 
            
            # 2. Cria as tabelas na DB 'dbexperience' que agora existe
            db.create_all()
            
            # 3. Execuções de inicialização normais
            criar_usuario_admin()
            migrar_banco_dados()
            atualizar_faturamento_sorteio()
            print("✅ Banco de dados PostgreSQL configurado com sucesso!")
            
        except Exception as e:
            # Captura qualquer erro (DB ou código) e informa
            print(f"❌ Erro ao conectar com PostgreSQL: {e}")
            print("🔧 Verifique a string de conexão, certificados e permissões.")
    
    # Configurações de HOST/PORTA para execução do servidor (Mantenha o seu código original abaixo)
    host = '0.0.0.0'
    
    if os.environ.get('SQUARECLOUD') or os.environ.get('PORT'):
        port = int(os.environ.get('PORT', 80))
        debug = False
        environment = "SquareCloud"
    else:
        port = 33053
        debug = True
        environment = "Desenvolvimento Local"
    
    print(f"🎯 R.O Experience 2025 - Servidor Iniciado!")
    print(f"📍 Host: {host}")
    print(f"🔧 Porta: {port}")
    print(f"🌐 Ambiente: {environment}")
    print(f"🐛 Debug: {debug}")
    print(f"🗄️ Banco: PostgreSQL")
    print("🚀 Aplicação rodando!")
    
    app.run(host=host, port=port, debug=debug)